#!/usr/bin/env python3
"""
cli.py — CLI interativo com suporte a banco de espera, criação interativa de linhas,
verificação da API AniList e migração.
"""

import os
import shlex
import sys
import time
import threading
import unicodedata
import random
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re

try:
    import requests
except Exception:
    print("Dependência 'requests' não encontrada. Instale com: pip install requests")
    sys.exit(1)

try:
    import colorama
    colorama.init(autoreset=True)
except Exception:
    colorama = None

def strip_ansi(text):
    """Remove códigos ANSI de uma string, inclusive sequências literais como [1;92m."""
    if not isinstance(text, str):
        return ""

    text = re.sub(r'\x1b\[[0-9;]*m', '', text)
    text = re.sub(r'\[(?:[0-9]{1,2}(?:;[0-9]{1,2})*)m', '', text)
    return text

ANSI_COLORS = {
    "black": "30",
    "red": "31",
    "green": "32",
    "yellow": "33",
    "blue": "34",
    "magenta": "35",
    "cyan": "36",
    "white": "37",
    "bright_black": "90",
    "bright_red": "91",
    "bright_green": "92",
    "bright_yellow": "93",
    "bright_blue": "94",
    "bright_magenta": "95",
    "bright_cyan": "96",
    "bright_white": "97"
}

ANSI_STYLES = {
    "bright": "1",
    "dim": "2",
    "normal": "22"
}


def supports_color():
    return hasattr(sys.stdout, "isatty") and sys.stdout.isatty()


def color_text(text, color=None, style=None):
    sanitized_text = strip_ansi(str(text))
    if not color or not supports_color():
        return sanitized_text
    codes = []
    if style and style in ANSI_STYLES:
        codes.append(ANSI_STYLES[style])
    if color in ANSI_COLORS:
        codes.append(ANSI_COLORS[color])
    if not codes:
        return sanitized_text
    return f"\033[{';'.join(codes)}m{sanitized_text}\033[0m"


def colored_prompt(text, color="green", style="bright"):
    return color_text(text, color=color, style=style)


def print_info(text, speed=0.0):
    typewriter_print(f"ℹ️  {text}", speed=speed, color="bright_blue", style="bright")


def print_success(text, speed=0.0):
    typewriter_print(f"✅ {text}", speed=speed, color="bright_green", style="bright")


def print_warning(text, speed=0.0):
    typewriter_print(f"⚠️  {text}", speed=speed, color="bright_yellow", style="bright")


def print_error(text, speed=0.0):
    typewriter_print(f"❌ {text}", speed=speed, color="bright_red", style="bright")


API_BASE = os.environ.get("API_BASE", "http://localhost:5000")
PROMPT_MAIN = color_text("┌─[menu]─", **{"color": "bright_cyan", "style": "bright"}) + color_text("$ ", **{"color": "bright_yellow", "style": "bright"})

OPINIAO_PRIORIDADES = {
    "Favorito": 0,
    "Muito Bom": 1,
    "Recomendo": 2,
    "Bom": 3,
    "Mediano": 4,
    "Ruim": 5,
    "Horrível": 6,
    "Horrivel": 6,
    "Não Vi": 7,
    "Nao Vi": 7
}

# ==================== ESTILOS DE COR ====================
STYLE = {
    "header": {"color": "bright_cyan", "style": "bright"},
    "command": {"color": "bright_yellow", "style": "bright"},
    "arg": {"color": "bright_green", "style": "bright"},
    "number": {"color": "bright_magenta", "style": "bright"},
    "error": {"color": "bright_red", "style": "bright"},
    "success": {"color": "bright_green", "style": "bright"},
    "info": {"color": "bright_blue", "style": "bright"},
    "warning": {"color": "bright_magenta", "style": "bright"},
    "dim": {"color": "white", "style": "dim"},
    "highlight": {"color": "bright_yellow", "style": "bright"},
}

# -------------------------
# Utilitários visuais/UX
# -------------------------
def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")

def spinner_worker(text, stop_event):
    symbols = ["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]
    i = 0
    while not stop_event.is_set():
        sys.stdout.write(color_text(f"\r⏳ {text} ", color="bright_cyan", style="bright") + color_text(symbols[i % len(symbols)], color="bright_yellow", style="bright"))
        sys.stdout.flush()
        time.sleep(0.12)
        i += 1
    sys.stdout.write("\r" + " " * (len(text) + 10) + "\r")
    sys.stdout.flush()

def with_minimum_spinner(fn, text="Processando", min_seconds=0.5, *args, **kwargs):
    stop_event = threading.Event()
    t = threading.Thread(target=spinner_worker, args=(text, stop_event), daemon=True)
    start = time.time()
    t.start()
    try:
        result = fn(*args, **kwargs)
    finally:
        elapsed = time.time() - start
        remaining = max(0, min_seconds - elapsed)
        if remaining > 0:
            time.sleep(remaining)
        stop_event.set()
        t.join()
    return result

def typewriter_print(text, speed=0.002, color=None, style=None):
    text = color_text(text, color=color, style=style)
    for ch in str(text):
        sys.stdout.write(ch)
        sys.stdout.flush()
        time.sleep(speed)
    sys.stdout.write("\n")
    sys.stdout.flush()

def fancy_header(lines, color="bright_cyan", border_char="═"):
    clear_screen()
    width = 80
    top = "╔" + border_char * (width - 2) + "╗"
    bottom = "╚" + border_char * (width - 2) + "╝"
    print(color_text(top, color=color, style="bright"))
    for ln in lines:
        ln_str = str(ln)
        if len(ln_str) > width - 4:
            ln_str = ln_str[:width - 7] + "..."
        padded = f"║ {ln_str.center(width - 4)} ║"
        print(color_text(padded, color=color, style="bright"))
    print(color_text(bottom, color=color, style="bright"))

# -------------------------
# Helpers: normalização
# -------------------------
def _strip_accents(s: str) -> str:
    if not isinstance(s, str):
        return ""
    nk = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in nk if not unicodedata.combining(ch))

def _norm(s: str) -> str:
    if not isinstance(s, str):
        return ""
    return _strip_accents(s).lower().strip()

def _norm_command_name(s: str) -> str:
    if not isinstance(s, str):
        return ""
    return _norm(s.replace("_", " ").replace("-", " "))

def search_items(itens, termo):
    termo_norm = _norm(termo)
    resultados = []
    for it in itens:
        if not isinstance(it, dict):
            continue
        nome = it.get("nome") or it.get("name") or ""
        if termo_norm in _norm(nome):
            resultados.append(it)
    return resultados


def enrich_created_line_from_anilist(line_id, nome, content_type, is_waiting=False):
    """Enriquece uma linha recém-criada com imagem, sinopse e sinônimos do AniList."""
    if is_waiting:
        return True, None

    media_type = "anime" if content_type in ["anime", "filme"] else "manga"
    try:
        image_resp = requests.get(
            f"{API_BASE.rstrip('/')}/search_image",
            params={"q": nome, "type": media_type},
            timeout=10,
        )
        if image_resp.ok:
            image_url = image_resp.json().get("image_url")
            if image_url:
                requests.put(
                    f"{API_BASE.rstrip('/')}/linhas/{line_id}/imagem",
                    json={"imagem_url": image_url},
                    timeout=8,
                )

        details_resp = requests.get(
            f"{API_BASE.rstrip('/')}/search_details",
            params={"q": nome, "type": media_type},
            timeout=10,
        )
        if details_resp.ok:
            details = details_resp.json()
            sinopse = details.get("sinopse")
            sinonimos = details.get("sinonimos")
            if sinopse is not None and sinonimos is not None:
                requests.put(
                    f"{API_BASE.rstrip('/')}/linhas/{line_id}/details",
                    json={"sinopse": sinopse, "sinonimos": sinonimos},
                    timeout=8,
                )
        return True, None
    except Exception as e:
        return False, str(e)


def _split_tags_field(tags_field):
    if not isinstance(tags_field, str):
        return []
    parts = [p.strip() for p in tags_field.split(",") if p.strip()]
    return [(_norm(p), p) for p in parts]

def tags_contains(item, tag_check):
    if not isinstance(item, dict):
        return False
    tags_field = item.get("tags") or ""
    norm_tag_check = _norm(tag_check)
    for norm, orig in _split_tags_field(tags_field):
        if norm == norm_tag_check:
            return True
    return False

# ============================================================
# SISTEMA DE TAGS LOCAL (do tagsSystem.js)

TAG_CATEGORIES = {
    "Romance": [
        "Romance", "Beijo", "Namoro", "Casamento", "Noivado",
        "Romance do bom", "Fez Filho(s)", "Gravidez"
    ],
    "Ação": [
        "Ação", "Poder", "Aventura", "Overpower", "Dungeon", "Mecha", "Demônio", "Monstros"
    ],
    "Fantasia": [
        "Magia", "Fantasia", "Sobrenatural", "Deuses", "Reencarnar", "Kemonomimi", "Medieval", "Goat", "Isekai", "MC Vilão"
    ],
    "Emocional": [
        "Drama", "Tristeza", "Vergonhoso", "Fofo"
    ],
    "Slice of Life": [
        "Slice of Life", "Vida Escolar", "Dormitorios", "Morar Juntos"
    ],
    "Tematico": [
        "Esporte", "Musical", "Terror", "Gore", "Comédia", "SciFi", "VR/Jogo", "System"
    ],
    "Gênero": [
        "Shounen", "Shoujo-ai", "Mahou Shoujo", "Yuri", "Gender bender"
    ],
    "Adulto": [
        "Ecchi", "Nudez", "Sexo", "Incesto", "NTR", "Harem", "Nudez Nippleless"
    ],
}


def get_all_tags_flat():
    """Retorna todas as tags em uma lista plana."""
    tags = []
    for category_tags in TAG_CATEGORIES.values():
        tags.extend(category_tags)
    return tags


def print_tags_table(tags, cols=3):
    """Exibe tags em 3 colunas com numeração colorida."""
    if not tags:
        print(color_text("Nenhuma tag encontrada.", **STYLE["dim"]))
        return
    
    sorted_tags = sorted(tags)
    rows = (len(sorted_tags) + cols - 1) // cols
    
    for r in range(rows):
        line = ""
        for c in range(cols):
            idx = r + c * rows
            if idx < len(sorted_tags):
                tag = sorted_tags[idx]
                display_tag = tag[:18] + ".." if len(tag) > 18 else tag
                num = color_text(f"{idx+1:3d}", **STYLE["number"])
                tag_colored = color_text(f"- {display_tag:20s}", color="bright_white", style="bright")
                line += f"{num} {tag_colored}"
            else:
                line += " " * 24
        print(line)

# -------------------------
# Help / documentação
# -------------------------
def print_help_main():
    fancy_header(["🚀 COMANDOS - MENU PRINCIPAL"], color="bright_cyan")
    commands = [
        ("show_lists", "Lista todas as listas disponíveis."),
        ("show_wait_lists", "Lista todas as listas no banco de espera."),
        ("create_list <nome>", "Cria uma nova lista no banco principal."),
        ("create_wait_list <nome>", "Cria uma nova lista no banco de espera."),
        ("open <id|nome>", "Abre uma lista pelo ID ou nome (principal)."),
        ("open_wait <id|nome>", "Abre uma lista do banco de espera."),
        ("delete_list <id|nome>", "Deleta uma lista (principal)."),
        ("migrate_wait <id>", "Migra seletivamente itens de uma lista de espera para o principal."),
        ("clear_wait", "Limpa todo o banco de espera (com confirmação)."),
        ("verify_api", "Verifica se a API do AniList está respondendo."),
        ("help | ?", "Mostra este help."),
        ("clear | cls", "Limpa a tela."),
        ("exit | quit", "Sai do CLI."),
        ("move", "Move itens entre listas do banco principal."),
    ]
    for cmd, desc in commands:
        cmd_col = color_text(cmd.ljust(22), **STYLE["command"])
        desc_col = color_text(desc, **STYLE["dim"])
        print(f"  {cmd_col}  {desc_col}")
    print()

def print_help_list():
    fancy_header(["📋 COMANDOS - LISTA ABERTA"], color="bright_green")
    commands = [
        ("show_lines [filtro]", "Exibe as linhas da lista."),
        ("show_tags", "Mostra todas as tags disponíveis."),
        ("search_<termo>", "Busca itens pelo nome."),
        ("open <nome>|<numero>", "Abre item por nome ou posição exibida."),
        ("show_<tag>", "Exibe itens da tag indicada."),
        ("show_anime|show_filme|...", "Filtra por conteúdo."),
        ("show_<status>", "Filtra por status."),
        ("sort_0-9 | sort_9-0", "Ordena por ID."),
        ("sort_a-z | sort_z-a", "Ordena por nome."),
        ("sort_rate [-r]", "Ordena por opinião."),
        ("next | prev", "Navega páginas."),
        ("<numero>", "Vai para a página indicada."),
        ("export_list [arquivo.xlsx]", "Exporta a exibição atual para XLSX."),
        ("create_line <nome>", "Cria uma nova linha interativamente na lista atual."),
        ("back | b", "Volta ao menu principal."),
        ("help | ?", "Mostra este help."),
        ("clear | cls", "Limpa a tela."),
        ("exit | quit", "Sai do CLI."),
    ]
    for cmd, desc in commands:
        cmd_col = color_text(cmd.ljust(24), **STYLE["command"])
        desc_col = color_text(desc, **STYLE["dim"])
        print(f"  {cmd_col}  {desc_col}")
    print()

def print_help_item():
    fancy_header(["📌 COMANDOS - ITEM ABERTO"], color="bright_magenta")
    commands = [
        ("next | n", "Abre o próximo item."),
        ("prev | p", "Abre o item anterior."),
        ("show_details", "Mostra detalhes completos."),
        ("edit <campo> <novo_valor>", "Edita um campo localmente."),
        ("edit", "Modo interativo de edição."),
        ("save", "Salva as alterações no servidor."),
        ("refresh", "Recarrega o item do servidor."),
        ("delete", "Exclui o item (com confirmação)."),
        ("check", "Atualiza o highlight."),
        ("back | b", "Volta para a lista."),
        ("help | ?", "Mostra este help."),
        ("clear | cls", "Limpa a tela."),
        ("exit | quit", "Sai do CLI."),
    ]
    for cmd, desc in commands:
        cmd_col = color_text(cmd.ljust(18), **STYLE["command"])
        desc_col = color_text(desc, **STYLE["dim"])
        print(f"  {cmd_col}  {desc_col}")
    print()

# -------------------------
# HTTP Requests
# -------------------------
def fetch_lists_request():
    url = f"{API_BASE.rstrip('/')}/listas"
    try:
        r = requests.get(url, timeout=6)
        if r.status_code >= 400:
            return None, f"Erro {r.status_code}: {r.text}"
        return r.json(), None
    except requests.exceptions.ConnectionError:
        return None, f"Erro de rede: não foi possível conectar ao servidor em {url}."
    except requests.exceptions.RequestException as e:
        return None, f"Erro de rede: {e}"

def fetch_wait_lists_request():
    url = f"{API_BASE.rstrip('/')}/wait/listas"
    try:
        r = requests.get(url, timeout=6)
        if r.status_code >= 400:
            return None, f"Erro {r.status_code}: {r.text}"
        return r.json(), None
    except Exception as e:
        return None, f"Erro: {e}"

def fetch_lines_request(list_id):
    url = f"{API_BASE.rstrip('/')}/linhas/{list_id}"
    try:
        r = requests.get(url, timeout=8)
        if r.status_code >= 400:
            return None, f"Erro {r.status_code}: {r.text}"
        return r.json(), None
    except Exception as e:
        return None, f"Erro: {e}"

def fetch_wait_lines_request(list_id):
    url = f"{API_BASE.rstrip('/')}/wait/linhas/{list_id}"
    try:
        r = requests.get(url, timeout=8)
        if r.status_code >= 400:
            return None, f"Erro {r.status_code}: {r.text}"
        return r.json(), None
    except Exception as e:
        return None, f"Erro: {e}"

def verify_anilist_api():
    import time
    from datetime import datetime
    
    url = "https://graphql.anilist.co"
    test_title = "Naruto"
    
    # Cabeçalho
    print("\n" + color_text("╔" + "═" * 78 + "╗", **STYLE["header"]))
    print(color_text("║" + " " * 20 + "🔍 VERIFICAÇÃO DA API ANILIST" + " " * 30 + "║", **STYLE["header"]))
    print(color_text("╚" + "═" * 78 + "╝", **STYLE["header"]))
    print(color_text(f"  📅 {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", **STYLE["dim"]))
    
    resultados = {
        "conectividade": False,
        "busca": False,
        "imagem": False,
        "sinopse": False,
        "sinonimos": False,
        "tempo_resposta": 0,
        "status_code": None,
        "erro": None,
        "detalhes": []
    }
    
    # 1) Teste de conectividade
    print(color_text("\n  📡 Testando conectividade...", **STYLE["info"]))
    try:
        start_time = time.time()
        
        test_query = """
        query {
          __typename
        }
        """
        
        r = requests.post(
            url,
            json={"query": test_query},
            timeout=5,
            headers={"Content-Type": "application/json"}
        )
        
        resultados["tempo_resposta"] = round((time.time() - start_time) * 1000, 2)
        resultados["status_code"] = r.status_code
        
        if r.status_code == 200:
            resultados["conectividade"] = True
            print(color_text(f"     ✅ Conectividade OK ({resultados['tempo_resposta']}ms)", **STYLE["success"]))
        else:
            print(color_text(f"     ⚠️ Resposta inesperada: HTTP {r.status_code}", **STYLE["warning"]))
            print(color_text(f"     📄 {r.text[:100]}", **STYLE["dim"]))
            resultados["erro"] = f"HTTP {r.status_code}"
            
    except requests.exceptions.ConnectionError:
        print(color_text("     ❌ FALHA - Sem conexão com a internet ou servidor bloqueado", **STYLE["error"]))
        resultados["erro"] = "Sem conexão com a internet"
    except requests.exceptions.Timeout:
        print(color_text("     ❌ FALHA - Tempo limite excedido (timeout)", **STYLE["error"]))
        resultados["erro"] = "Timeout"
    except Exception as e:
        print(color_text(f"     ❌ FALHA - {str(e)}", **STYLE["error"]))
        resultados["erro"] = str(e)
    
    # Se não houver conectividade, já exibe o relatório
    if not resultados["conectividade"]:
        print("\n" + color_text("╔" + "═" * 78 + "╗", **STYLE["header"]))
        print(color_text("║" + " " * 25 + "📊 RELATÓRIO DA API" + " " * 33 + "║", **STYLE["header"]))
        print(color_text("╚" + "═" * 78 + "╝", **STYLE["header"]))
        
        print(color_text("\n  🔴 STATUS GERAL: ❌ API INDISPONÍVEL", **STYLE["error"]))
        print(color_text(f"  🔴 Motivo: {resultados['erro']}", **STYLE["error"]))
        print(color_text("  🔴 Testes passados: 0/5", **STYLE["error"]))
        print(color_text(f"  🔴 Tempo de resposta: {resultados['tempo_resposta']}ms", **STYLE["error"]))
        
        print(color_text("\n  📋 Detalhamento:", **STYLE["info"]))
        print(color_text("     ❌ Conectividade com a API", **STYLE["error"]))
        print(color_text("     ❌ Busca por título", **STYLE["error"]))
        print(color_text("     ❌ Retorno de imagem", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinopse", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinônimos", **STYLE["error"]))
        
        print(color_text("\n  💡 RECOMENDAÇÃO:", **STYLE["warning"]))
        print(color_text("     ⚠️ A API do AniList está inacessível no momento.", **STYLE["warning"]))
        print(color_text("     📌 Use o banco de ESPERA para adicionar itens.", **STYLE["dim"]))
        print(color_text("     📌 Depois execute 'migrate_wait' quando a API voltar.", **STYLE["dim"]))
        print(color_text("     📌 Verifique sua conexão com a internet.", **STYLE["dim"]))
        print("\n" + color_text("═" * 80, **STYLE["dim"]))
        return False, resultados
    
    # 2) Query de busca
    query = """
    query($search: String) {
      Page(page: 1, perPage: 3) {
        media(search: $search, type: ANIME) {
          title { romaji english native }
          coverImage { large extraLarge }
          description
          synonyms
          status
          episodes
          averageScore
        }
      }
    }
    """
    
    print(color_text(f"\n  🔎 Buscando por: '{test_title}'...", **STYLE["info"]))
    
    try:
        start_time = time.time()
        response = requests.post(
            url,
            json={"query": query, "variables": {"search": test_title}},
            timeout=10,
            headers={"Content-Type": "application/json"}
        )
        response_time = round((time.time() - start_time) * 1000, 2)
        
        if response.status_code != 200:
            print(color_text(f"     ❌ Erro na busca: HTTP {response.status_code}", **STYLE["error"]))
            print(color_text(f"     📄 Resposta: {response.text[:150]}...", **STYLE["dim"]))
            resultados["erro"] = f"HTTP {response.status_code} na busca"
            resultados["status_code"] = response.status_code
            
            print("\n" + color_text("╔" + "═" * 78 + "╗", **STYLE["header"]))
            print(color_text("║" + " " * 25 + "📊 RELATÓRIO DA API" + " " * 33 + "║", **STYLE["header"]))
            print(color_text("╚" + "═" * 78 + "╝", **STYLE["header"]))
            
            print(color_text("\n  🔴 STATUS GERAL: ⚠️ API RESPONDEU MAS FALHOU NA BUSCA", **STYLE["warning"]))
            print(color_text(f"  🔴 Motivo: {resultados['erro']}", **STYLE["warning"]))
            print(color_text("  🔴 Testes passados: 1/5", **STYLE["warning"]))
            print(color_text(f"  🔴 Tempo de resposta: {resultados['tempo_resposta']}ms", **STYLE["warning"]))
            
            print(color_text("\n  📋 Detalhamento:", **STYLE["info"]))
            print(color_text("     ✅ Conectividade com a API", **STYLE["success"]))
            print(color_text("     ❌ Busca por título", **STYLE["error"]))
            print(color_text("     ❌ Retorno de imagem", **STYLE["error"]))
            print(color_text("     ❌ Retorno de sinopse", **STYLE["error"]))
            print(color_text("     ❌ Retorno de sinônimos", **STYLE["error"]))
            
            print(color_text("\n  💡 RECOMENDAÇÃO:", **STYLE["warning"]))
            print(color_text("     ⚠️ A API está respondendo mas a busca falhou.", **STYLE["warning"]))
            print(color_text("     📌 Pode ser um problema temporário. Tente novamente em alguns minutos.", **STYLE["dim"]))
            print(color_text("     📌 Use o banco de ESPERA para adicionar itens agora.", **STYLE["dim"]))
            print("\n" + color_text("═" * 80, **STYLE["dim"]))
            return False, resultados
        
        data = response.json()
        resultados["busca"] = True
        print(color_text(f"     ✅ Busca realizada com sucesso ({response_time}ms)", **STYLE["success"]))
        
    except requests.exceptions.Timeout:
        print(color_text("     ❌ Timeout na busca", **STYLE["error"]))
        resultados["erro"] = "Timeout na busca"
        
        print("\n" + color_text("╔" + "═" * 78 + "╗", **STYLE["header"]))
        print(color_text("║" + " " * 25 + "📊 RELATÓRIO DA API" + " " * 33 + "║", **STYLE["header"]))
        print(color_text("╚" + "═" * 78 + "╝", **STYLE["header"]))
        
        print(color_text("\n  🔴 STATUS GERAL: ⚠️ TIMEOUT NA BUSCA", **STYLE["warning"]))
        print(color_text("  🔴 Testes passados: 1/5", **STYLE["warning"]))
        print(color_text(f"  🔴 Tempo de resposta: {resultados['tempo_resposta']}ms", **STYLE["warning"]))
        
        print(color_text("\n  📋 Detalhamento:", **STYLE["info"]))
        print(color_text("     ✅ Conectividade com a API", **STYLE["success"]))
        print(color_text("     ❌ Busca por título (timeout)", **STYLE["error"]))
        print(color_text("     ❌ Retorno de imagem", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinopse", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinônimos", **STYLE["error"]))
        
        print(color_text("\n  💡 RECOMENDAÇÃO:", **STYLE["warning"]))
        print(color_text("     ⚠️ A API está lenta. Tente novamente mais tarde.", **STYLE["warning"]))
        print(color_text("     📌 Use o banco de ESPERA para adicionar itens agora.", **STYLE["dim"]))
        print("\n" + color_text("═" * 80, **STYLE["dim"]))
        return False, resultados
        
    except Exception as e:
        print(color_text(f"     ❌ Erro: {e}", **STYLE["error"]))
        resultados["erro"] = str(e)
        
        print("\n" + color_text("╔" + "═" * 78 + "╗", **STYLE["header"]))
        print(color_text("║" + " " * 25 + "📊 RELATÓRIO DA API" + " " * 33 + "║", **STYLE["header"]))
        print(color_text("╚" + "═" * 78 + "╝", **STYLE["header"]))
        
        print(color_text("\n  🔴 STATUS GERAL: ⚠️ ERRO NA BUSCA", **STYLE["warning"]))
        print(color_text(f"  🔴 Motivo: {resultados['erro']}", **STYLE["warning"]))
        print(color_text("  🔴 Testes passados: 1/5", **STYLE["warning"]))
        
        print(color_text("\n  📋 Detalhamento:", **STYLE["info"]))
        print(color_text("     ✅ Conectividade com a API", **STYLE["success"]))
        print(color_text("     ❌ Busca por título (erro)", **STYLE["error"]))
        print(color_text("     ❌ Retorno de imagem", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinopse", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinônimos", **STYLE["error"]))
        
        print(color_text("\n  💡 RECOMENDAÇÃO:", **STYLE["warning"]))
        print(color_text("     ⚠️ Ocorreu um erro inesperado.", **STYLE["warning"]))
        print(color_text("     📌 Use o banco de ESPERA para adicionar itens.", **STYLE["dim"]))
        print("\n" + color_text("═" * 80, **STYLE["dim"]))
        return False, resultados
    
    # 3) Analisar resultados da busca
    media = data.get("data", {}).get("Page", {}).get("media", [])
    
    if not media:
        print(color_text("     ⚠️ Nenhum resultado encontrado", **STYLE["warning"]))
        resultados["erro"] = "Nenhum resultado encontrado"
        
        print("\n" + color_text("╔" + "═" * 78 + "╗", **STYLE["header"]))
        print(color_text("║" + " " * 25 + "📊 RELATÓRIO DA API" + " " * 33 + "║", **STYLE["header"]))
        print(color_text("╚" + "═" * 78 + "╝", **STYLE["header"]))
        
        print(color_text("\n  🔴 STATUS GERAL: ⚠️ BUSCA SEM RESULTADOS", **STYLE["warning"]))
        print(color_text("  🔴 Testes passados: 2/5", **STYLE["warning"]))
        print(color_text(f"  🔴 Tempo de resposta: {resultados['tempo_resposta']}ms", **STYLE["warning"]))
        
        print(color_text("\n  📋 Detalhamento:", **STYLE["info"]))
        print(color_text("     ✅ Conectividade com a API", **STYLE["success"]))
        print(color_text("     ✅ Busca por título (mas sem resultados)", **STYLE["success"]))
        print(color_text("     ❌ Retorno de imagem", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinopse", **STYLE["error"]))
        print(color_text("     ❌ Retorno de sinônimos", **STYLE["error"]))
        
        print(color_text("\n  💡 RECOMENDAÇÃO:", **STYLE["warning"]))
        print(color_text("     ⚠️ A API respondeu mas não encontrou o título de teste.", **STYLE["warning"]))
        print(color_text("     📌 Pode ser um problema nos dados da API.", **STYLE["dim"]))
        print(color_text("     📌 Use o banco de ESPERA para adicionar itens.", **STYLE["dim"]))
        print("\n" + color_text("═" * 80, **STYLE["dim"]))
        return False, resultados
    
    # 4) Analisar dados do primeiro resultado
    primeiro = media[0]
    titulos = primeiro.get("title", {})
    romaji = titulos.get("romaji", "N/A")
    english = titulos.get("english", "")
    native = titulos.get("native", "")
    
    print(color_text(f"\n  📋 Resultado encontrado:", **STYLE["info"]))
    print(color_text(f"     📖 Título: {romaji}", color="bright_white", style="bright"))
    if english:
        print(color_text(f"     🌐 Inglês: {english}", **STYLE["dim"]))
    if native:
        print(color_text(f"     🇯🇵 Nativo: {native}", **STYLE["dim"]))
    
    # 4.1) Imagem
    cover = primeiro.get("coverImage", {})
    large = cover.get("large", "")
    extra = cover.get("extraLarge", "")
    
    if large:
        resultados["imagem"] = True
        print(color_text(f"     🖼️ Imagem: ✅ Disponível (large)", **STYLE["success"]))
        try:
            img_check = requests.head(large, timeout=3)
            if img_check.status_code == 200:
                print(color_text(f"        📸 URL válida (HTTP {img_check.status_code})", **STYLE["dim"]))
            else:
                print(color_text(f"        ⚠️ URL retornou HTTP {img_check.status_code}", **STYLE["warning"]))
        except:
            print(color_text(f"        ⚠️ Não foi possível verificar a URL", **STYLE["warning"]))
    elif extra:
        resultados["imagem"] = True
        print(color_text(f"     🖼️ Imagem: ✅ Disponível (extraLarge)", **STYLE["success"]))
    else:
        print(color_text(f"     🖼️ Imagem: ❌ NÃO DISPONÍVEL", **STYLE["error"]))
        resultados["detalhes"].append("Imagem não disponível")
    
    # 4.2) Sinopse
    desc = primeiro.get("description", "")
    if desc and len(desc) > 50:
        resultados["sinopse"] = True
        print(color_text(f"     📝 Sinopse: ✅ Disponível ({len(desc)} caracteres)", **STYLE["success"]))
        preview = desc[:120].replace("\n", " ").strip()
        print(color_text(f"        📄 {preview}...", **STYLE["dim"]))
    elif desc:
        resultados["sinopse"] = True
        print(color_text(f"     📝 Sinopse: ✅ Disponível (curta, {len(desc)} caracteres)", **STYLE["success"]))
    else:
        print(color_text(f"     📝 Sinopse: ❌ NÃO DISPONÍVEL", **STYLE["error"]))
        resultados["detalhes"].append("Sinopse não disponível")
    
    # 4.3) Sinônimos
    synonyms = primeiro.get("synonyms", [])
    if synonyms and len(synonyms) > 0:
        resultados["sinonimos"] = True
        print(color_text(f"     🔤 Sinônimos: ✅ Disponível ({len(synonyms)} sinônimo(s))", **STYLE["success"]))
        display = ", ".join(synonyms[:3])
        if len(synonyms) > 3:
            display += f" (+{len(synonyms)-3} mais)"
        print(color_text(f"        📌 {display}", **STYLE["dim"]))
    else:
        print(color_text(f"     🔤 Sinônimos: ❌ NÃO DISPONÍVEL", **STYLE["error"]))
        resultados["detalhes"].append("Sinônimos não disponíveis")
    
    # 4.4) Informações extras
    status = primeiro.get("status", "N/A")
    episodes = primeiro.get("episodes", "N/A")
    score = primeiro.get("averageScore", "N/A")
    
    print(color_text(f"\n  📊 Informações adicionais:", **STYLE["info"]))
    print(color_text(f"     📌 Status: {status}", **STYLE["dim"]))
    print(color_text(f"     📌 Episódios: {episodes}", **STYLE["dim"]))
    print(color_text(f"     📌 Score médio: {score}/100", **STYLE["dim"]))
    
    # 5) Sumário final
    print("\n" + color_text("╔" + "═" * 78 + "╗", **STYLE["header"]))
    print(color_text("║" + " " * 28 + "📊 RELATÓRIO DA API" + " " * 33 + "║", **STYLE["header"]))
    print(color_text("╚" + "═" * 78 + "╝", **STYLE["header"]))
    
    total_tests = 5
    passed = sum([
        resultados["conectividade"],
        resultados["busca"],
        resultados["imagem"],
        resultados["sinopse"],
        resultados["sinonimos"]
    ])
    
    if passed == 5:
        status_emoji = "🟢"
        status_text = "COMPLETAMENTE FUNCIONAL"
        status_color = "bright_green"
    elif passed >= 3:
        status_emoji = "🟡"
        status_text = "PARCIALMENTE FUNCIONAL"
        status_color = "bright_yellow"
    else:
        status_emoji = "🔴"
        status_text = "COM PROBLEMAS"
        status_color = "bright_red"
    
    print(color_text(f"\n  {status_emoji} STATUS GERAL: {status_text}", color=status_color, style="bright"))
    print(color_text(f"  📊 Testes passados: {passed}/{total_tests}", **STYLE["info"]))
    print(color_text(f"  ⏱️ Tempo de resposta: {resultados['tempo_resposta']}ms", **STYLE["info"]))
    
    if resultados.get("status_code"):
        print(color_text(f"  📡 HTTP Status: {resultados['status_code']}", **STYLE["info"]))
    
    print(color_text("\n  📋 Detalhamento:", **STYLE["info"]))
    print(color_text(f"     {'✅' if resultados['conectividade'] else '❌'} Conectividade com a API", **STYLE["success"] if resultados['conectividade'] else STYLE["error"]))
    print(color_text(f"     {'✅' if resultados['busca'] else '❌'} Busca por título", **STYLE["success"] if resultados['busca'] else STYLE["error"]))
    print(color_text(f"     {'✅' if resultados['imagem'] else '❌'} Retorno de imagem", **STYLE["success"] if resultados['imagem'] else STYLE["error"]))
    print(color_text(f"     {'✅' if resultados['sinopse'] else '❌'} Retorno de sinopse", **STYLE["success"] if resultados['sinopse'] else STYLE["error"]))
    print(color_text(f"     {'✅' if resultados['sinonimos'] else '❌'} Retorno de sinônimos", **STYLE["success"] if resultados['sinonimos'] else STYLE["error"]))
    
    if resultados["detalhes"]:
        print(color_text("\n  📌 Observações:", **STYLE["warning"]))
        for det in resultados["detalhes"]:
            print(color_text(f"     ⚠️ {det}", **STYLE["warning"]))
    
    # 6) Recomendações
    print(color_text("\n  💡 RECOMENDAÇÃO:", **STYLE["highlight"]))
    if passed == 5:
        print(color_text("     ✅ Tudo funcionando perfeitamente!", **STYLE["success"]))
        print(color_text("     ✅ Pode usar o banco PRINCIPAL normalmente.", **STYLE["success"]))
        print(color_text("     ✅ As imagens, sinopse e sinônimos estão disponíveis.", **STYLE["success"]))
    elif passed >= 3:
        print(color_text("     ⚠️ API está funcionando parcialmente.", **STYLE["warning"]))
        print(color_text("     📌 Alguns dados podem não estar disponíveis.", **STYLE["dim"]))
        print(color_text("     📌 Use o banco PRINCIPAL com cautela.", **STYLE["dim"]))
        if not resultados["imagem"]:
            print(color_text("     📌 Imagens não disponíveis. Considere adicionar manualmente.", **STYLE["dim"]))
        if not resultados["sinopse"]:
            print(color_text("     📌 Sinopse não disponível.", **STYLE["dim"]))
        if not resultados["sinonimos"]:
            print(color_text("     📌 Sinônimos não disponíveis.", **STYLE["dim"]))
    else:
        print(color_text("     🔴 API com problemas ou indisponível.", **STYLE["error"]))
        print(color_text("     📌 Use o banco de ESPERA para adicionar itens.", **STYLE["dim"]))
        print(color_text("     📌 Execute 'migrate_wait' quando a API normalizar.", **STYLE["dim"]))
        if resultados.get("erro"):
            print(color_text(f"     📌 Motivo: {resultados['erro']}", **STYLE["dim"]))
    
    print("\n" + color_text("═" * 80, **STYLE["dim"]))
    
    return passed >= 3, resultados


# -------------------------
# Paginação
# -------------------------
class PaginatedDisplay:
    def __init__(self, items, title, items_per_page=None):
        self.items = items or []
        self.title = title
        self.items_per_page = None if items_per_page is None else self._clamp(items_per_page)
        self.pagination_confirmed = self.items_per_page is not None
        self.current_page = 1
        self.total_pages = 1

    def _clamp(self, n):
        try:
            n = int(n)
        except Exception:
            return 5
        if n < 1:
            return 1
        return min(15, n)

    def _recalc_pages(self):
        ipp = self.items_per_page or 5
        self.total_pages = max(1, (len(self.items) + ipp - 1) // ipp)
        self.current_page = max(1, min(self.current_page, self.total_pages))

    def ask_items_per_page(self):
        prompt_default = 5
        while True:
            try:
                raw = input(color_text(f"Quantos itens por página? [1-15] (enter={prompt_default}): ", **STYLE["info"])).strip()
            except (KeyboardInterrupt, EOFError):
                print()
                raw = ""
            if raw == "":
                chosen = prompt_default
                break
            try:
                chosen = int(raw)
            except Exception:
                print(color_text("Entrada inválida — informe um número entre 1 e 15.", **STYLE["error"]))
                continue
            if 1 <= chosen <= 15:
                break
            else:
                print(color_text("Valor fora do intervalo. Informe entre 1 e 15.", **STYLE["error"]))
        return chosen

    def render_page(self, page_num=None):
        if not self.pagination_confirmed:
            chosen = self.ask_items_per_page()
            self.items_per_page = self._clamp(chosen)
            self.pagination_confirmed = True
            self._recalc_pages()

        if page_num:
            try:
                self.current_page = max(1, int(page_num))
            except Exception:
                self.current_page = 1

        self._recalc_pages()
        start_idx = (self.current_page - 1) * self.items_per_page
        end_idx = start_idx + self.items_per_page
        page_items = self.items[start_idx:end_idx]

        fancy_header([
            "════════════════════════════════════════════════════════════════════════════════",
            f"         {self.title}",
            f"         Página {self.current_page} de {self.total_pages}  (itens/pg: {self.items_per_page})",
            "════════════════════════════════════════════════════════════════════════════════",
        ], color="bright_cyan")

        if not page_items:
            print(color_text("(nenhum item encontrado nesta seleção)", color="bright_yellow", style="bright"))
        else:
            for idx, item in enumerate(page_items, start=start_idx + 1):
                self._pretty_print_item(idx, item)

        print(color_text("-" * 80, **STYLE["dim"]))
        print(color_text(self._compact_page_display(), **STYLE["highlight"]))
        print(color_text("Navegue com 'next', 'prev' ou digite o número da página. Use 'back' para voltar.", **STYLE["dim"]))
        print(color_text(f"Total: {len(self.items)} itens. (mostrando {self.items_per_page} por página)", **STYLE["dim"]))
        print(color_text("=" * 80, **STYLE["dim"]))

    def _compact_page_display(self):
        if self.total_pages <= 12:
            parts = []
            for p in range(1, self.total_pages + 1):
                if p == self.current_page:
                    parts.append(color_text(f"[{p}]", **STYLE["highlight"]))
                else:
                    parts.append(color_text(str(p), **STYLE["dim"]))
            return " ".join(parts)
        parts = []
        parts.append(color_text("1", **STYLE["highlight"]) if self.current_page == 1 else color_text("1", **STYLE["dim"]))
        if self.current_page > 4:
            parts.append(color_text("...", **STYLE["dim"]))
        start = max(2, self.current_page - 2)
        end = min(self.total_pages - 1, self.current_page + 2)
        for p in range(start, end + 1):
            if p == self.current_page:
                parts.append(color_text(f"[{p}]", **STYLE["highlight"]))
            else:
                parts.append(color_text(str(p), **STYLE["dim"]))
        if self.current_page < self.total_pages - 3:
            parts.append(color_text("...", **STYLE["dim"]))
        if self.current_page == self.total_pages:
            parts.append(color_text(f"[{self.total_pages}]", **STYLE["highlight"]))
        else:
            parts.append(color_text(str(self.total_pages), **STYLE["dim"]))
        return " ".join(parts)

    def _pretty_print_item(self, idx, item):
        if not isinstance(item, dict):
            print(f"{idx}. {item}")
            return
        
        WIDTH = 80
        MAX_NAME_LEN = 50
        
        nome = item.get("nome") or item.get("name") or str(item.get("id", "N/A"))
        migrated = item.get("migrated", 0)
        
        # ====== TRATA O NOME (corta se passar do limite) ======
        if len(nome) > MAX_NAME_LEN:
            nome_display = nome[:MAX_NAME_LEN - 3] + "..."
        else:
            nome_display = nome
        
        # ====== CORES ======
        idx_str = color_text(f"{idx:3d}.", **STYLE["number"])
        
        conteudo = item.get("conteudo", "").lower()
        if conteudo in ["anime", "filme"]:
            name_color = "bright_cyan"
        elif conteudo in ["manga", "manhwa", "webtoon"]:
            name_color = "bright_green"
        else:
            name_color = "white"
        
        nome_str = color_text(nome_display, color=name_color, style="bright")
        prefix = color_text("[M] ", **STYLE["warning"]) if migrated else ""
        
        # ====== STATUS (SÓ ELE) ======
        status = item.get("status")
        status_str = ""
        if status:
            status_color = "bright_yellow"
            if status.lower() in ["concluido", "finished"]:
                status_color = "bright_green"
            elif status.lower() in ["dropado", "cancelado"]:
                status_color = "bright_red"
            status_str = color_text(f"[{status}]", color=status_color, style="bright")
        
        # ====== MONTAGEM COM ALINHAMENTO ======
        # Parte esquerda (sem cores para cálculo)
        left_raw = f"{idx:3d}. {prefix}{nome_display}" if migrated else f"{idx:3d}. {nome_display}"
        
        # Parte direita (sem cores para cálculo)
        right_raw = f"  {status}" if status else ""
        
        # Padding
        padding = max(1, WIDTH - len(left_raw) - len(right_raw))
        
        # Linha final
        if status:
            line = f"{idx_str} {prefix}{nome_str}" + " " * padding + status_str
        else:
            line = f"{idx_str} {prefix}{nome_str}"
        
        sys.stdout.write(line + "\n")
        sys.stdout.flush()

    def handle_command(self, cmd, args):
        cmd = str(cmd).strip().lower()
        if cmd == "next" and self.current_page < self.total_pages:
            self.render_page(self.current_page + 1)
            return True
        if cmd == "prev" and self.current_page > 1:
            self.render_page(self.current_page - 1)
            return True
        if cmd.isdigit():
            page_num = int(cmd)
            if 1 <= page_num <= self.total_pages:
                self.render_page(page_num)
                return True
        return False

    def apply_sort(self, method):
        if not self.items:
            return "Nenhum item para ordenar."
        method = method.lower()
        try:
            if method == "0-9":
                def key_id(x):
                    if isinstance(x, dict):
                        v = x.get("id", x.get("Id", None))
                    else:
                        v = x
                    try:
                        return int(v)
                    except Exception:
                        return _norm(str(v))
                self.items = sorted(self.items, key=key_id)
                self.current_page = 1
                return "Ordenado por id crescente (0-9)."
            if method == "9-0":
                def key_id(x):
                    if isinstance(x, dict):
                        v = x.get("id", x.get("Id", None))
                    else:
                        v = x
                    try:
                        return int(v)
                    except Exception:
                        return _norm(str(v))
                self.items = sorted(self.items, key=key_id, reverse=True)
                self.current_page = 1
                return "Ordenado por id decrescente (9-0)."
            if method == "a-z":
                def key_name(x):
                    if isinstance(x, dict):
                        return _norm(x.get("nome") or x.get("name") or str(x.get("id","")))
                    return _norm(str(x))
                self.items = sorted(self.items, key=key_name)
                self.current_page = 1
                return "Ordenado A→Z."
            if method == "z-a":
                def key_name(x):
                    if isinstance(x, dict):
                        return _norm(x.get("nome") or x.get("name") or str(x.get("id","")))
                    return _norm(str(x))
                self.items = sorted(self.items, key=key_name, reverse=True)
                self.current_page = 1
                return "Ordenado Z→A."
            if method in ("rate", "rate -r"):
                opiniao_order = ["Favorito", "Muito Bom", "Recomendo", "Bom", "Mediano", "Ruim", "Horrivel", "Horrível", "Não Vi", "Nao Vi"]
                def get_priority(item):
                    tags_field = item.get("tags") if isinstance(item, dict) else ""
                    tags_norm = [t for (t, orig) in _split_tags_field(tags_field)]
                    def has_tag(t): return _norm(t) in tags_norm
                    has_relation = any(has_tag(x) for x in ("namoro", "casamento", "noivado"))
                    is_bestlove = has_tag("goat") and has_tag("beijo") and has_tag("romance do bom") and has_relation
                    if is_bestlove:
                        return 0
                    is_love = (has_tag("beijo") and has_tag("romance do bom") and has_relation)
                    if is_love and has_tag("goat"):
                        return 1
                    if has_tag("goat"):
                        return 2
                    opiniao_raw = item.get("opiniao") if isinstance(item, dict) else ""
                    opiniao_norm = _norm(opiniao_raw)
                    if is_love and opiniao_norm == _norm("Favorito"):
                        return 3
                    for idx, op in enumerate(opiniao_order):
                        if _norm(op) == opiniao_norm:
                            return 4 + idx
                    return 99
                def key_rate(x):
                    try:
                        return (get_priority(x), _norm(x.get("nome") or x.get("name") or str(x.get("id",""))))
                    except Exception:
                        return (99, _norm(str(x)))
                reverse = (method == "rate -r")
                self.items = sorted(self.items, key=key_rate, reverse=reverse)
                self.current_page = 1
                return "Ordenado por 'rate' (reverse)." if reverse else "Ordenado por 'rate'."
            return "Método de ordenação desconhecido."
        except Exception as e:
            return f"Erro ao ordenar: {e}"

# -------------------------
# Funções auxiliares de filtro e parser
# -------------------------
def extrair_tags_dos_itens(itens):
    todas_tags = set()
    for item in itens:
        if isinstance(item, dict) and item.get("tags"):
            tags = [tag.strip() for tag in item["tags"].split(",") if tag.strip()]
            todas_tags.update(tags)
    return sorted(todas_tags)

def filtrar_por_tag(itens, tag_procurada):
    tag_procurada = tag_procurada.lower()
    resultados = []
    for item in itens:
        if isinstance(item, dict) and item.get("tags"):
            tags_item = [tag.strip().lower() for tag in item["tags"].split(",")]
            if tag_procurada in tags_item:
                resultados.append(item)
    return resultados

def filtrar_por_status(itens, status_procurado):
    status_procurado = status_procurado.lower()
    return [it for it in itens if isinstance(it, dict) and it.get("status","").lower() == status_procurado]

def filtrar_por_opiniao(itens, opiniao_procurada):
    opiniao_procurada = opiniao_procurada.lower()
    return [it for it in itens if isinstance(it, dict) and it.get("opiniao","").lower() == opiniao_procurada]

def filtrar_por_comparacao_opiniao(itens, expressao):
    operadores = [">=", "<=", ">", "<", "="]
    operador = None
    opiniao_ref = expressao
    for op in operadores:
        if expressao.startswith(op):
            operador = op
            opiniao_ref = expressao[len(op):].strip()
            break
    if opiniao_ref not in OPINIAO_PRIORIDADES:
        return []
    valor_ref = OPINIAO_PRIORIDADES[opiniao_ref]
    resultados = []
    for item in itens:
        if not isinstance(item, dict) or "opiniao" not in item:
            continue
        opiniao_item = item["opiniao"]
        if opiniao_item not in OPINIAO_PRIORIDADES:
            continue
        valor_item = OPINIAO_PRIORIDADES[opiniao_item]
        if operador == ">":
            if valor_item < valor_ref:
                resultados.append(item)
        elif operador == ">=":
            if valor_item <= valor_ref:
                resultados.append(item)
        elif operador == "<":
            if valor_item > valor_ref:
                resultados.append(item)
        elif operador == "<=":
            if valor_item >= valor_ref:
                resultados.append(item)
        elif operador == "=" or operador is None:
            if valor_item == valor_ref:
                resultados.append(item)
    return resultados

def parse_search_expression(expr_str):
    import shlex
    if not expr_str:
        return {
            "required_tags": set(),
            "excluded_tags": set(),
            "statuses": set(),
            "content_types": set(),
            "opiniao_cmp": None
        }
    CONTENT_TYPES = {"anime", "filme", "manga", "manhwa", "webtoon"}
    STATUS_SET = {"concluido", "vendo", "cancelado", "dropado", "lendo", "assistir", "conheço", "assistindo", "finished"}
    tokens = shlex.split(expr_str)
    required_tags = set()
    excluded_tags = set()
    content_types = set()
    statuses = set()
    opiniao_cmp = None
    opiniao_map = {k.casefold(): k for k in OPINIAO_PRIORIDADES.keys()}
    i = 0
    while i < len(tokens):
        tk = tokens[i]
        i += 1
        for op in (">=", "<=", ">", "<", "="):
            if tk.startswith(op):
                val = tk[len(op):].strip()
                if not val and i < len(tokens):
                    val = tokens[i]; i += 1
                key = val.casefold()
                if key in opiniao_map:
                    opiniao_cmp = op + opiniao_map[key]
                else:
                    opiniao_cmp = op + val
                tk = None
                break
        if tk is None:
            continue
        if tk.startswith("+") or tk.startswith("-"):
            sign = tk[0]
            val = tk[1:].strip()
            if "|" in val:
                parts = [p.strip().casefold() for p in val.split("|") if p.strip()]
                for p in parts:
                    if p in CONTENT_TYPES:
                        content_types.add(p)
                    elif p in STATUS_SET:
                        statuses.add(p)
                    elif p in opiniao_map:
                        opiniao_cmp = "=" + opiniao_map[p]
                continue
            plain = val.replace("_", " ").strip()
            low = plain.casefold()
            if low in CONTENT_TYPES:
                if sign == "+":
                    content_types.add(low)
                else:
                    excluded_tags.add(f"__conteudo__:{low}")
                continue
            if low in STATUS_SET:
                if sign == "+":
                    statuses.add(low)
                continue
            if low in opiniao_map:
                if sign == "+":
                    opiniao_cmp = "=" + opiniao_map[low]
                continue
            if sign == "+":
                required_tags.add(low)
            else:
                excluded_tags.add(low)
            continue
        plain = tk.replace("_", " ").strip()
        low = plain.casefold()
        if "|" in tk:
            parts = [p.strip().casefold() for p in tk.split("|") if p.strip()]
            for p in parts:
                if p in CONTENT_TYPES:
                    content_types.add(p)
                else:
                    required_tags.add(p)
            continue
        if low in CONTENT_TYPES:
            content_types.add(low)
            continue
        if low in STATUS_SET:
            statuses.add(low)
            continue
        if low in opiniao_map:
            opiniao_cmp = "=" + opiniao_map[low]
            continue
        required_tags.add(low)
    return {
        "required_tags": required_tags,
        "excluded_tags": excluded_tags,
        "statuses": statuses,
        "content_types": content_types,
        "opiniao_cmp": opiniao_cmp
    }

# -------------------------
# Contextos
# -------------------------
class OpenListContext:
    def __init__(self, list_obj, is_waiting=False):
        self.list_obj = list_obj
        self.id = str(list_obj.get("id") or "")
        self.name = list_obj.get("nome") or self.id or "lista"
        self.is_waiting = is_waiting
        self.lines = []
        self.current_display = None

    def fetch_and_cache_lines(self):
        if self.is_waiting:
            lines, err = with_minimum_spinner(
                lambda: fetch_wait_lines_request(self.id),
                text=f"Buscando linhas da lista de espera '{self.name}'...",
                min_seconds=0.6
            )
        else:
            lines, err = with_minimum_spinner(
                lambda: fetch_lines_request(self.id),
                text=f"Buscando linhas da lista '{self.name}'...",
                min_seconds=0.6
            )
        if err:
            return False, err
        if not isinstance(lines, list):
            return False, "Resposta inesperada."
        def keyfn(it):
            if isinstance(it, dict):
                return (it.get("nome") or it.get("name") or "").casefold()
            return str(it).casefold()
        self.lines = sorted(lines, key=keyfn)
        return True, None

    def open_item_by_index(self, one_based_index):
        items = None
        if getattr(self, "current_display", None) and getattr(self.current_display, "items", None):
            items = self.current_display.items
        else:
            items = self.lines
        try:
            idx = int(one_based_index)
        except Exception:
            return None, "Índice inválido."
        if idx < 1 or idx > len(items):
            return None, f"Índice fora do intervalo (1..{len(items)})"
        item = items[idx - 1]
        return ItemContext(self, item, idx), None

    def open_item_by_name(self, nome):
        if not isinstance(nome, str) or not nome.strip():
            return None, "Nome inválido."
        nome_normalizado = _norm_command_name(nome)
        exact_match = None
        partial_matches = []
        for idx, item in enumerate(self.lines, start=1):
            if not isinstance(item, dict):
                continue
            item_name = item.get("nome") or item.get("name") or ""
            item_norm = _norm_command_name(item_name)
            if item_norm == nome_normalizado:
                exact_match = (idx, item)
                break
            if nome_normalizado in item_norm:
                partial_matches.append((idx, item))
        if exact_match:
            idx, item = exact_match
            return ItemContext(self, item, idx), None
        if len(partial_matches) == 1:
            idx, item = partial_matches[0]
            return ItemContext(self, item, idx), None
        if len(partial_matches) > 1:
            nomes = [f"{idx}:{item.get('nome') or item.get('name')}" for idx, item in partial_matches[:5]]
            return None, f"Vários itens correspondem a '{nome}': {', '.join(nomes)}"
        return None, f"Nenhum item encontrado com nome '{nome}'."

    def search_items_by_name(self, termo):
        itens = search_items(self.lines, termo)
        titulo = f"RESULTADOS DE BUSCA: \"{termo}\" - {self.name}"
        self.current_display = PaginatedDisplay(itens, titulo, items_per_page=None)
        self.current_display.render_page()

    def show_lines(self, filtro_expresao=None):
        itens = list(self.lines)
        parsed = parse_search_expression(filtro_expresao) if filtro_expresao else None
        if parsed:
            if parsed["content_types"]:
                allowed = set(parsed["content_types"])
                itens = [it for it in itens if isinstance(it, dict) and it.get("conteudo") and it.get("conteudo").casefold() in allowed]
            if parsed["opiniao_cmp"]:
                itens = filtrar_por_comparacao_opiniao(itens, parsed["opiniao_cmp"])
            if parsed["statuses"]:
                sts = set(parsed["statuses"])
                itens = [it for it in itens if isinstance(it, dict) and it.get("status","").casefold() in sts]
            if parsed["excluded_tags"]:
                excl = set(parsed["excluded_tags"])
                def has_excluded(it):
                    if not isinstance(it, dict):
                        return False
                    tags = [t.strip().casefold() for t in (it.get("tags") or "").split(",") if t.strip()]
                    for ex in excl:
                        if ex.startswith("__conteudo__:"):
                            need = ex.split(":",1)[1]
                            if (it.get("conteudo") or "").casefold() == need:
                                return True
                        if ex in tags:
                            return True
                    return False
                itens = [it for it in itens if not has_excluded(it)]
            if parsed["required_tags"]:
                req = set(parsed["required_tags"])
                def has_all_required(it):
                    if not isinstance(it, dict):
                        return False
                    tags = [t.strip().casefold() for t in (it.get("tags") or "").split(",") if t.strip()]
                    return req.issubset(set(tags))
                itens = [it for it in itens if has_all_required(it)]
        titulo = f"LINHAS DA LISTA: {self.name}"
        if filtro_expresao:
            titulo += f" [Filtro: {filtro_expresao}]"
        self.current_display = PaginatedDisplay(itens, titulo, items_per_page=None)
        self.current_display.render_page()

    def show_tags(self):
        tags = extrair_tags_dos_itens(self.lines)
        self.current_display = PaginatedDisplay(tags, f"TAGS DISPONÍVEIS - {self.name}", items_per_page=None)
        self.current_display.render_page()

    def show_por_tag(self, tag):
        itens = filtrar_por_tag(self.lines, tag)
        self.current_display = PaginatedDisplay(itens, f"ITENS COM TAG: {tag} - {self.name}", items_per_page=None)
        self.current_display.render_page()

    def show_por_status(self, status):
        itens = filtrar_por_status(self.lines, status)
        self.current_display = PaginatedDisplay(itens, f"ITENS {status.capitalize()} - {self.name}", items_per_page=None)
        self.current_display.render_page()

    def show_por_opiniao(self, opiniao):
        itens = filtrar_por_opiniao(self.lines, opiniao)
        self.current_display = PaginatedDisplay(itens, f"ITENS COM OPINIÃO: {opiniao.capitalize()} - {self.name}", items_per_page=None)
        self.current_display.render_page()

    def show_por_content(self, content_type_or_iter):
        if isinstance(content_type_or_iter, str):
            types = {content_type_or_iter.casefold()}
        else:
            types = {t.casefold() for t in content_type_or_iter}
        itens = [it for it in self.lines if isinstance(it, dict) and (it.get("conteudo") or "").casefold() in types]
        caps = ", ".join([t.capitalize() for t in types])
        self.current_display = PaginatedDisplay(itens, f"ITENS ({caps}) - {self.name}", items_per_page=None)
        self.current_display.render_page()

    def export_current_display(self, filename_arg=None):
        if not getattr(self, "current_display", None):
            return False, "Nenhuma exibição ativa. Use 'show_lines' primeiro."
        items = list(self.current_display.items or [])
        if not items:
            return False, "Nenhum item visível para exportar."
        default_fname = f"{self.name}.xlsx"
        if filename_arg:
            filename = filename_arg.strip()
        else:
            raw = input(color_text(f"Nome do arquivo (enter={default_fname}): ", **STYLE["info"])).strip()
            filename = raw or default_fname
        if not filename.lower().endswith(".xlsx"):
            filename = filename + ".xlsx"

        def ask_opt(prompt, default=True):
            yn = "Y/n" if default else "y/N"
            raw = input(color_text(f"{prompt} [{yn}]: ", **STYLE["info"])).strip().lower()
            if raw == "":
                return default
            return raw[0] == "y"

        opts = {
            "id": ask_opt("Incluir ID?", True),
            "nome": ask_opt("Incluir Nome?", True),
            "sinonimos": ask_opt("Incluir Sinônimos?", True),
            "tag": ask_opt("Incluir Tag?", True),
            "opiniao": ask_opt("Incluir Opinião?", True),
            "episodio": ask_opt("Incluir Ep/Cap?", True),
            "status": ask_opt("Incluir Status?", True),
            "sinopse": ask_opt("Incluir Sinopse?", True),
            "conteudo": ask_opt("Incluir Conteúdo?", True),
            "image": ask_opt("Incluir Imagem (URL)?", True),
        }

        color_map = {}
        used = set()
        def rand_color():
            while True:
                r = random.randint(0, 200)
                g = random.randint(0, 200)
                b = random.randint(0, 200)
                hexc = "{:02X}{:02X}{:02X}".format(r, g, b)
                if hexc not in used:
                    used.add(hexc)
                    return "FF" + hexc
        for it in items:
            key = str(it.get("id", str(id(it)))) if isinstance(it, dict) else str(it)
            color_map[key] = rand_color()

        header_keys = []
        if opts["id"]: header_keys.append("ID")
        if opts["nome"]: header_keys.append("Nome")
        if opts["sinonimos"]: header_keys.append("Sinonimos")
        if opts["tag"]: header_keys.append("Tag")
        if opts["opiniao"]: header_keys.append("Opinião")
        if opts["episodio"]: header_keys.append("Ep/Cap")
        if opts["status"]: header_keys.append("Status")
        if opts["sinopse"]: header_keys.append("Sinopse")
        if opts["conteudo"]: header_keys.append("Conteudo")
        if opts["image"]: header_keys.append("Imagem")

        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        ws.append(header_keys)

        total = len(items)
        current = 0

        for it in items:
            key = str(it.get("id", str(id(it)))) if isinstance(it, dict) else str(it)
            bg = color_map.get(key, "FFDDDDDD")
            tags_field = ""
            if isinstance(it, dict):
                tags_field = it.get("tags") or ""
            if tags_field:
                tags = [t.strip() for t in tags_field.split(",") if t.strip()]
                if not tags:
                    tags = [""]
            else:
                tags = [""]

            for tag in tags:
                row = []
                if opts["id"]:
                    row.append(it.get("id") if isinstance(it, dict) else "")
                if opts["nome"]:
                    row.append(it.get("nome") if isinstance(it, dict) else str(it))
                if opts["sinonimos"]:
                    val = ""
                    if isinstance(it, dict):
                        s = it.get("sinonimos")
                        if isinstance(s, (list, tuple)):
                            val = "; ".join(s)
                        else:
                            val = s or ""
                    row.append(val)
                if opts["tag"]:
                    row.append(tag)
                if opts["opiniao"]:
                    row.append(it.get("opiniao") if isinstance(it, dict) else "")
                if opts["episodio"]:
                    row.append(it.get("episodio") if isinstance(it, dict) else "")
                if opts["status"]:
                    row.append(it.get("status") if isinstance(it, dict) else "")
                if opts["sinopse"]:
                    row.append(it.get("sinopse") if isinstance(it, dict) else "")
                if opts["conteudo"]:
                    row.append(it.get("conteudo") if isinstance(it, dict) else "")
                if opts["image"]:
                    row.append(it.get("imagem_url") if isinstance(it, dict) else "")

                ws.append(row)
                last_row_idx = ws.max_row
                fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                for col_idx in range(1, len(header_keys) + 1):
                    cell = ws.cell(row=last_row_idx, column=col_idx)
                    cell.fill = fill

            current += 1
            pct = int((current / total) * 100)
            sys.stdout.write(color_text(f"\rExportando... {current}/{total} ({pct}%)", **STYLE["info"]))
            sys.stdout.flush()

        wb.save(filename)
        print()
        return True, f"Arquivo salvo: {os.path.abspath(filename)}"

    # ============================================================
    # Criação interativa de linha (CORRIGIDA E ESTILIZADA)
    # ============================================================

    def interactive_create_line(self, nome, is_waiting=False):
        """
        Realiza o fluxo interativo de criação de uma nova linha.
        USANDO TAGS DO ARQUIVO LOCAL (não busca do banco).
        """
        
        # 1) Usar tags do arquivo local
        all_tags = get_all_tags_flat()
        
        # 2) Exibir tags em colunas (5 colunas)
        fancy_header([f"📝 CRIANDO NOVA LINHA: {nome}"], color="bright_yellow")
        print(color_text("\n📋 TAGS DISPONÍVEIS (do sistema):", **STYLE["info"]))
        print(color_text("-" * 80, **STYLE["dim"]))
        
        # Pergunta tags
        print_tags_table(all_tags)
        print(color_text("-" * 80, **STYLE["dim"]))
        tag_choice = input(color_text("Quais tags essa linha vai ter? (digite os números separados por vírgula, ou deixe em branco): ", **STYLE["info"])).strip()
        selected_tags = []
        if tag_choice:
            sorted_tags = sorted(all_tags)
            for part in tag_choice.split(','):
                part = part.strip()
                if part.isdigit():
                    idx = int(part) - 1
                    if 0 <= idx < len(sorted_tags):
                        selected_tags.append(sorted_tags[idx])
                    else:
                        print(color_text(f"⚠️ Número {part} inválido (fora do range). Ignorando.", **STYLE["warning"]))
            if selected_tags:
                print(color_text(f"✅ Tags selecionadas: {', '.join(selected_tags)}", **STYLE["success"]))
            else:
                print(color_text("ℹ️ Nenhuma tag válida selecionada.", **STYLE["dim"]))
        
        tags_str = ", ".join(selected_tags) if selected_tags else ""
        input(color_text("\nPressione ENTER para continuar...", **STYLE["dim"]))

        # 3) Tipo de mídia
        fancy_header([f"🎬 TIPO DE MÍDIA para '{nome}'"], color="bright_cyan")
        media_options = [
            ("Anime", "anime"),
            ("Filme", "filme"),
            ("Manga", "manga"),
            ("Manhwa", "manhwa"),
            ("Webtoon", "webtoon")
        ]
        for i, (label, _) in enumerate(media_options, 1):
            print(color_text(f"{i} - {label}", color="bright_white", style="bright"))
        while True:
            choice = input(color_text("Escolha o tipo (número): ", **STYLE["info"])).strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(media_options):
                    conteudo = media_options[idx][1]
                    break
            print(color_text("Opção inválida. Tente novamente.", **STYLE["error"]))

        # 4) Status (baseado no tipo)
        fancy_header([f"📊 STATUS para '{nome}'"], color="bright_magenta")
        if conteudo in ["anime", "filme"]:
            status_options = [
                ("Assistindo", "assistindo"),
                ("Concluído", "concluido"),
                ("Assistir", "assistir"),
                ("Cancelado", "cancelado"),
                ("Dropado", "dropado")
            ]
        else:
            status_options = [
                ("Lendo", "lendo"),
                ("Concluído", "concluido"),
                ("Ler", "ler"),
                ("Cancelado", "cancelado"),
                ("Dropado", "dropado")
            ]
        for i, (label, _) in enumerate(status_options, 1):
            print(color_text(f"{i} - {label}", color="bright_white", style="bright"))
        while True:
            choice = input(color_text("Escolha o status (número): ", **STYLE["info"])).strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(status_options):
                    status = status_options[idx][1]
                    break
            print(color_text("Opção inválida. Tente novamente.", **STYLE["error"]))

        # 5) Episódio/Capítulo
        fancy_header([f"📖 EPISÓDIO/CAPÍTULO para '{nome}'"], color="bright_green")
        episodio_input = input(color_text("Em qual episódio/capítulo você parou? (deixe em branco se não aplicável): ", **STYLE["info"])).strip()
        episodio = None
        if episodio_input:
            try:
                episodio = int(episodio_input)
            except ValueError:
                try:
                    episodio = float(episodio_input)
                except ValueError:
                    print(color_text("⚠️ Valor inválido. Será salvo como vazio.", **STYLE["warning"]))
                    episodio = None

        # 6) Opinião
        fancy_header([f"⭐ OPINIÃO sobre '{nome}'"], color="bright_yellow")
        opiniao_options = [
            ("Favorito", "Favorito"),
            ("Muito Bom", "Muito Bom"),
            ("Recomendo", "Recomendo"),
            ("Bom", "Bom"),
            ("Mediano", "Mediano"),
            ("Ruim", "Ruim"),
            ("Horrível", "Horrível"),
            ("Não Vi", "Não Vi")
        ]
        for i, (label, _) in enumerate(opiniao_options, 1):
            print(color_text(f"{i} - {label}", color="bright_white", style="bright"))
        while True:
            choice = input(color_text("Escolha a opinião (número): ", **STYLE["info"])).strip()
            if choice.isdigit():
                idx = int(choice) - 1
                if 0 <= idx < len(opiniao_options):
                    opiniao = opiniao_options[idx][1]
                    break
            print(color_text("Opção inválida. Tente novamente.", **STYLE["error"]))

        # 7) Confirmar e enviar
        fancy_header(["📋 RESUMO DA CRIAÇÃO"], color="bright_cyan")
        print(color_text(f"📌 Nome: {nome}", color="bright_white", style="bright"))
        print(color_text(f"🏷️ Tags: {tags_str or '(nenhuma)'}", **STYLE["dim"]))
        print(color_text(f"📺 Tipo: {conteudo}", color="bright_cyan", style="bright"))
        print(color_text(f"📊 Status: {status}", **STYLE["info"]))
        print(color_text(f"📖 Episódio/Cap: {episodio if episodio is not None else '(não informado)'}", **STYLE["dim"]))
        print(color_text(f"⭐ Opinião: {opiniao}", color="bright_yellow", style="bright"))
        print(color_text("=" * 80, **STYLE["dim"]))
        confirm = input(color_text("Criar esta linha? (s/N): ", **STYLE["highlight"])).strip().lower()
        if confirm != 's':
            print(color_text("❌ Criação cancelada.", **STYLE["error"]))
            return

        # 8) Enviar para o servidor
        payload = {
            "lista_id": self.id,
            "nome": nome,
            "tags": tags_str,
            "conteudo": conteudo,
            "status": status,
            "episodio": episodio,
            "opiniao": opiniao
        }
        
        if is_waiting:
            url = f"{API_BASE.rstrip('/')}/wait/linhas"
        else:
            url = f"{API_BASE.rstrip('/')}/linhas"

        try:
            print(color_text("\n⏳ Enviando para o servidor...", **STYLE["info"]))
            r = requests.post(url, json=payload, timeout=8)
            if r.status_code >= 400:
                print(color_text(f"❌ Erro ao criar linha: {r.status_code} - {r.text}", **STYLE["error"]))
            else:
                print(color_text("✅ Linha criada com sucesso!", **STYLE["success"]))
                if not is_waiting:
                    created = r.json()
                    line_id = created.get("id")
                    if line_id:
                        print(color_text("⏳ Buscando dados do AniList para imagem, sinopse e sinônimos...", **STYLE["info"]))
                        ok, err = enrich_created_line_from_anilist(
                            line_id,
                            nome,
                            conteudo,
                            is_waiting=is_waiting,
                        )
                        if not ok:
                            print(color_text(f"⚠️ Não foi possível enriquecer a linha no momento: {err}", **STYLE["warning"]))
                self.fetch_and_cache_lines()
                self.show_lines()
        except Exception as e:
            print(color_text(f"❌ Erro de rede: {e}", **STYLE["error"]))

    def _print_tags_table(self, tags):
        """Exibe tags em 3 colunas com numeração."""
        if not tags:
            print(color_text("Nenhuma tag encontrada no sistema.", **STYLE["dim"]))
            return
        cols = 3
        rows = (len(tags) + cols - 1) // cols
        for r in range(rows):
            line = ""
            for c in range(cols):
                idx = r + c * rows
                if idx < len(tags):
                    tag = tags[idx]
                    display_tag = tag[:18] + ".." if len(tag) > 18 else tag
                    num = color_text(f"{idx+1:2d}", **STYLE["number"])
                    line += f"{num} - {display_tag:20s}"
                else:
                    line += " " * 24
            print(line)

class ItemContext:
    def __init__(self, parent_ctx, item, idx_in_view):
        self.parent = parent_ctx
        self.item = dict(item) if isinstance(item, dict) else {"nome": str(item)}
        self.index_in_view = int(idx_in_view)
        self.name = str(self.item.get("id") or self.item.get("nome") or f"item{self.index_in_view}")
        self.modified = False

    def show_details(self):
        i = self.item
        clear_screen()
        
        # Cores suaves
        label_color = "bright_black"  # Cinza claro para os rótulos
        value_color = "white"         # Branco para os valores
        header_color = "bright_cyan"  # Ciano suave para o cabeçalho
        
        # ====== CABEÇALHO ======
        print(color_text("═" * 80, color="bright_black", style="dim"))
        print(color_text(f"  📌 ITEM #{self.index_in_view}  |  ID: {i.get('id')}", color=header_color, style="bright"))
        print(color_text("═" * 80, color="bright_black", style="dim"))
        
        # ====== LINHA 1: Nome ======
        print(f"  {color_text('Nome:', color=label_color)}  {color_text(i.get('nome') or 'N/A', color=value_color, style='bright')}")
        
        # ====== LINHA 2: Conteúdo + Status + Opinião ======
        conteudo = i.get('conteudo') or 'N/A'
        status = i.get('status') or 'N/A'
        opiniao = i.get('opiniao') or 'N/A'
        
        # Cor do conteúdo
        conteudo_color = "bright_cyan" if conteudo.lower() in ["anime", "filme"] else "bright_green"
        
        # Cor do status
        status_color = "bright_green" if status.lower() in ["concluido", "finished"] else "bright_yellow"
        if status.lower() in ["dropado", "cancelado"]:
            status_color = "bright_red"
        
        # Cor da opinião
        opiniao_color = "white"
        if opiniao == "Favorito":
            opiniao_color = "bright_magenta"
        elif opiniao in ["Muito Bom", "Recomendo"]:
            opiniao_color = "bright_green"
        
        print(f"  {color_text('Conteúdo:', color=label_color)}  {color_text(conteudo, color=conteudo_color, style='bright')}  |  {color_text('Status:', color=label_color)}  {color_text(status, color=status_color, style='bright')}  |  {color_text('Opinião:', color=label_color)}  {color_text(opiniao, color=opiniao_color, style='bright')}")
        
        # ====== LINHA 3: Episódio/Capítulo ======
        episodio = i.get('episodio')
        episodio_str = str(episodio) if episodio is not None else 'N/A'
        print(f"  {color_text('Episódio/Capítulo:', color=label_color)}  {color_text(episodio_str, color=value_color, style='bright')}")
        
        # ====== LINHA 4: Tags ======
        tags = i.get('tags') or ''
        if tags:
            # Colore cada tag individualmente
            tag_list = [t.strip() for t in tags.split(',') if t.strip()]
            colored_tags = []
            for tag in tag_list:
                # Cores aleatórias mas suaves para tags
                tag_colors = ["bright_cyan", "bright_green", "bright_yellow", "bright_magenta", "bright_blue"]
                tag_color = tag_colors[hash(tag) % len(tag_colors)]
                colored_tags.append(color_text(tag, color=tag_color, style="bright"))
            tags_display = "  ".join(colored_tags)
        else:
            tags_display = color_text("(nenhuma)", color="bright_black")
        
        print(f"  {color_text('Tags:', color=label_color)}  {tags_display}")
        
        # ====== LINHA 5: Sinônimos ======
        sinonimos = i.get('sinonimos')
        if sinonimos:
            if isinstance(sinonimos, list):
                sinonimos_str = "  ".join(sinonimos)
            else:
                sinonimos_str = str(sinonimos)
        else:
            sinonimos_str = color_text("(nenhum)", color="bright_black")
        print(f"  {color_text('Sinônimos:', color=label_color)}  {color_text(sinonimos_str, color=value_color)}")
        
        # ====== LINHA 6: Imagem URL ======
        imagem = i.get('imagem_url') or i.get('image') or ''
        if imagem:
            # Mostra a URL encurtada visualmente mas mantém o texto completo
            display_url = imagem
            print(f"  {color_text('Imagem:', color=label_color)}  {color_text(display_url, color='bright_blue')}")
        else:
            print(f"  {color_text('Imagem:', color=label_color)}  {color_text('(nenhuma)', color='bright_black')}")
        
        # ====== LINHA 7: Sinopse ======
        sinopse = i.get('sinopse') or ''
        print(f"  {color_text('Sinopse:', color=label_color)}")
        if sinopse:
            # Quebra a sinopse em linhas de 76 caracteres
            lines = []
            for j in range(0, len(sinopse), 76):
                lines.append(f"    {sinopse[j:j+76]}")
            for line in lines:
                print(color_text(line, color=value_color))
        else:
            print(f"    {color_text('(sem sinopse)', color='bright_black')}")
        
        # ====== RODAPÉ ======
        print(color_text("═" * 80, color="bright_black", style="dim"))
        print(color_text("  Comandos: next | prev | edit | save | refresh | delete | check | back", color="bright_black", style="dim"))
        print(color_text("═" * 80, color="bright_black", style="dim"))

    def edit_field(self, field, new_value):
        field = field.strip().lower()
        if not field:
            return "Campo inválido."
        aliases = {
            "ep": "episodio",
            "episode": "episodio",
            "image": "imagem_url",
            "imagem": "imagem_url",
            "sinonyms": "sinonimos",
            "sinonym": "sinonimos",
            "name": "nome",
        }
        field = aliases.get(field, field)
        self.item[field] = new_value
        self.modified = True
        return f"Campo '{field}' atualizado localmente."

    def interactive_edit(self):
        editable = ["nome","conteudo","status","episodio","opiniao","tags","sinopse","imagem_url","sinonimos"]
        print(color_text("Modo interativo — deixe em branco para manter o valor atual.", **STYLE["info"]))
        for f in editable:
            cur = self.item.get(f, "")
            raw = input(color_text(f"{f} (atual: {cur}) => ", **STYLE["info"])).rstrip("\n")
            if raw != "":
                if f == "sinonimos" and ";" in raw:
                    self.item[f] = [s.strip() for s in raw.split(";") if s.strip()]
                else:
                    self.item[f] = raw
                self.modified = True
        return "Edição local concluída."

    def save(self):
        if "id" not in self.item:
            return False, "Item sem ID, não é possível salvar."
        tags_value = self.item.get("tags")
        if isinstance(tags_value, (list, tuple)):
            tags_value = ", ".join(str(x).strip() for x in tags_value if x is not None)
        episodio_value = self.item.get("episodio")
        if isinstance(episodio_value, str):
            episodio_value = episodio_value.strip()
            try:
                episodio_value = int(episodio_value)
            except ValueError:
                try:
                    episodio_value = float(episodio_value)
                except ValueError:
                    pass
        payload = {
            "nome": self.item.get("nome"),
            "conteudo": self.item.get("conteudo"),
            "status": self.item.get("status"),
            "episodio": episodio_value,
            "opiniao": self.item.get("opiniao"),
            "tags": tags_value,
            "sinopse": self.item.get("sinopse"),
            "imagem_url": self.item.get("imagem_url"),
            "sinonimos": self.item.get("sinonimos")
        }
        payload = {k: v for k, v in payload.items() if v is not None}
        url = f"{API_BASE.rstrip('/')}/linhas/{self.item['id']}"
        try:
            r = requests.put(url, json=payload, timeout=8)
            if r.status_code >= 400:
                return False, f"Erro {r.status_code}: {r.text}"
            self.modified = False
            return True, f"Salvo com sucesso: {self.item['id']}"
        except Exception as e:
            return False, f"Erro de rede: {e}"

    def refresh(self):
        if "id" not in self.item:
            return False, "Item sem ID."
        url = f"{API_BASE.rstrip('/')}/linhas/{self.item['id']}"
        try:
            r = requests.get(url, timeout=6)
            if r.status_code >= 400:
                return False, f"Erro {r.status_code}: {r.text}"
            new = r.json()
            if isinstance(new, dict):
                self.item = new
                return True, "Dados atualizados do servidor."
            return False, "Resposta inesperada do servidor."
        except Exception as e:
            return False, f"Erro de rede: {e}"

    def delete(self):
        if "id" not in self.item:
            return False, "Item sem ID."
        url = f"{API_BASE.rstrip('/')}/linhas/{self.item['id']}"
        try:
            r = requests.delete(url, timeout=6)
            if r.status_code >= 400:
                return False, f"Erro {r.status_code}: {r.text}"
            return True, "Excluído com sucesso."
        except Exception as e:
            return False, f"Erro de rede: {e}"

    def check(self):
        if "id" not in self.item:
            return False, "Item sem ID."
        url = f"{API_BASE.rstrip('/')}/highlighted/{self.item['id']}"
        try:
            r = requests.post(url, timeout=8)
            if r.status_code >= 400:
                return False, f"Erro {r.status_code}: {r.text}"
            try:
                payload = r.json()
                if isinstance(payload, dict) and "mensagem" in payload:
                    self.item["last_highlight"] = payload.get("last_highlight", self.item.get("last_highlight"))
            except Exception:
                pass
            return True, "Highlight atualizado com sucesso."
        except Exception as e:
            return False, f"Erro de rede: {e}"

    def open_adjacent(self, offset):
        target_idx = self.index_in_view - 1 + offset
        items = self.parent.current_display.items if getattr(self.parent, "current_display", None) else self.parent.lines
        if target_idx < 0 or target_idx >= len(items):
            return None
        return ItemContext(self.parent, items[target_idx], target_idx + 1)

# -------------------------
# Funções de abertura e criação de listas
# -------------------------
def cmd_open_list(raw_name, is_waiting=False):
    key = raw_name.strip()
    key_norm = _norm_command_name(key)
    if is_waiting:
        (listas, err) = with_minimum_spinner(
            lambda: fetch_wait_lists_request(),
            text=f"Procurando lista de espera '{key}'...",
            min_seconds=0.6
        )
    else:
        (listas, err) = with_minimum_spinner(
            lambda: fetch_lists_request(),
            text=f"Procurando lista '{key}'...",
            min_seconds=0.6
        )
    if err:
        fancy_header([f"❌ Erro: {err}"], color="bright_red")
        return None
    match = None
    partial_matches = []
    for item in listas or []:
        if not isinstance(item, dict):
            continue
        if str(item.get("id")) == key:
            match = item
            break
        nome = item.get("nome") or ""
        nome_norm = _norm_command_name(nome)
        if nome_norm == key_norm:
            match = item
            break
        if key_norm and key_norm in nome_norm:
            partial_matches.append(item)
    if not match:
        if len(partial_matches) == 1:
            match = partial_matches[0]
        elif len(partial_matches) > 1:
            nomes = [item.get('nome') or str(item.get('id')) for item in partial_matches[:5]]
            fancy_header([f"⚠️ Várias listas correspondem a '{key}': {', '.join(nomes)}"], color="bright_yellow")
            return None
    if not match:
        fancy_header([f"❌ Não encontrei a lista '{key}'.", "Use 'show_lists' para ver todas."], color="bright_red")
        return None
    ctx = OpenListContext(match, is_waiting=is_waiting)
    ok, fetch_err = ctx.fetch_and_cache_lines()
    if not ok:
        fancy_header([f"❌ Erro ao carregar linhas: {fetch_err}"], color="bright_red")
        return None
    fancy_header([f"✅ LISTA '{match.get('nome') or match.get('id')}' ABERTA" +
                  (" (ESPERA)" if is_waiting else "")], color="bright_green")
    return ctx

def cmd_delete_list(raw_name):
    key = raw_name.strip()
    key_norm = _norm_command_name(key)
    (listas, err) = with_minimum_spinner(lambda: fetch_lists_request(), text=f"Procurando lista '{key}'...", min_seconds=0.6)
    if err:
        print_error(f"Erro ao buscar listas: {err}")
        return False
    match = None
    partial_matches = []
    for item in listas or []:
        if not isinstance(item, dict):
            continue
        id_str = str(item.get("id") or "")
        name = item.get("nome") or item.get("name") or ""
        if key_norm == _norm_command_name(id_str) or key_norm == _norm_command_name(name):
            match = item
            break
        if key_norm in _norm_command_name(name) or key_norm in _norm_command_name(id_str):
            partial_matches.append(item)
    if not match:
        if len(partial_matches) == 1:
            match = partial_matches[0]
        elif partial_matches:
            print_warning("Várias correspondências encontradas:")
            for it in partial_matches:
                print(color_text(f" - {it.get('id')} : {it.get('nome')}", **STYLE["dim"]))
            print(color_text("Seja mais específico.", **STYLE["info"]))
            return False
        else:
            print_error("Lista não encontrada.")
            return False
    confirm = input(color_text(f"⚠️ Tem certeza que deseja deletar a lista '{match.get('nome')}' (ID {match.get('id')})? [y/N]: ", **STYLE["warning"])).strip().lower()
    if confirm != 'y':
        print_info("Operação cancelada.")
        return False
    url = f"{API_BASE.rstrip('/')}/listas/{match.get('id')}"
    try:
        r = requests.delete(url, timeout=8)
        if r.status_code >= 400:
            print_error(f"Erro ao deletar lista: HTTP {r.status_code} - {getattr(r, 'text', '')}")
            return False
        print_success(f"Lista '{match.get('nome')}' deletada com sucesso.")
        return True
    except requests.exceptions.RequestException as e:
        print_error(f"Erro de requisição: {e}")
        return False

def cmd_create_list(nome, is_waiting=False):
    """Cria uma nova lista (principal ou espera)."""
    url = f"{API_BASE.rstrip('/')}/wait/listas" if is_waiting else f"{API_BASE.rstrip('/')}/listas"
    try:
        r = requests.post(url, json={"nome": nome}, timeout=6)
        if r.status_code >= 400:
            print_error(f"Erro ao criar lista: {r.status_code} - {r.text}")
        else:
            data = r.json()
            print_success(f"Lista '{nome}' criada com sucesso (ID: {data.get('id')})")
    except Exception as e:
        print_error(f"Erro de rede: {e}")

def cmd_migrate_wait(wait_list_id=None):
    """
    Migração seletiva: exibe linhas da lista de espera, usuário escolhe quais
    pela posição (número) e para qual lista principal.
    """
    if not wait_list_id:
        print_error("Uso: migrate_wait <id_lista_espera>")
        return

    # 1. Buscar lista de espera
    listas_espera, err = fetch_wait_lists_request()
    if err:
        print_error(f"Erro ao buscar listas de espera: {err}")
        return
    wait_list = next((l for l in listas_espera if str(l.get('id')) == str(wait_list_id)), None)
    if not wait_list:
        print_error(f"Lista de espera com ID {wait_list_id} não encontrada.")
        return

    # 2. Buscar linhas da lista de espera
    linhas, err = fetch_wait_lines_request(wait_list_id)
    if err:
        print_error(f"Erro ao buscar linhas: {err}")
        return
    if not linhas:
        print_info("Esta lista de espera está vazia.")
        return

    linhas_ordenadas = sorted(linhas, key=lambda x: x.get('nome', '').casefold())

    # 3. Exibir linhas numeradas com indicador de migração
    fancy_header([f"📋 LISTA DE ESPERA: {wait_list['nome']} (ID {wait_list_id})"], color="bright_cyan")
    for idx, linha in enumerate(linhas_ordenadas, start=1):
        migrado = linha.get('migrated', 0)
        marcador = color_text("[M] ", **STYLE["warning"]) if migrado else "    "
        id_colored = color_text(f"(ID:{linha['id']})", **STYLE["dim"])
        print(f"{color_text(str(idx).rjust(3), **STYLE['number'])}. {marcador}{color_text(linha['nome'], color='bright_white', style='bright')} {id_colored}")
    print(color_text("-" * 80, **STYLE["dim"]))

    # 4. Perguntar quais linhas migrar (por POSIÇÃO)
    while True:
        raw = input(color_text("Quais linhas você quer migrar? (digite os números de posição separados por vírgula, ex: 1,3,5): ", **STYLE["info"])).strip()
        if not raw:
            print_info("Nenhum número informado. Operação cancelada.")
            return
        indices = []
        for part in raw.split(','):
            part = part.strip()
            if part.isdigit():
                indices.append(int(part))
        if not indices:
            print_error("Números inválidos. Tente novamente.")
            continue
        invalidos = [i for i in indices if i < 1 or i > len(linhas_ordenadas)]
        if invalidos:
            print_error(f"Números fora do intervalo (1..{len(linhas_ordenadas)}): {', '.join(map(str, invalidos))}")
            continue
        linhas_selecionadas = [linhas_ordenadas[i-1] for i in indices]
        ja_migrados = [l for l in linhas_selecionadas if l.get('migrated', 0) == 1]
        if ja_migrados:
            nomes = [l['nome'] for l in ja_migrados]
            print_warning(f"As seguintes linhas já foram migradas anteriormente: {', '.join(nomes)}")
            continuar = input(color_text("Deseja continuar apenas com as não migradas? (s/N): ", **STYLE["warning"])).strip().lower()
            if continuar != 's':
                continue
            linhas_selecionadas = [l for l in linhas_selecionadas if l.get('migrated', 0) == 0]
        if not linhas_selecionadas:
            print_info("Nenhuma linha válida para migrar. Cancelando.")
            return
        break

    # 5. Buscar listas principais
    listas_principais, err = fetch_lists_request()
    if err:
        print_error(f"Erro ao buscar listas principais: {err}")
        return
    if not listas_principais:
        print_info("Nenhuma lista principal disponível. Crie uma primeiro.")
        return

    # 6. Exibir listas principais numeradas
    fancy_header(["📋 LISTAS PRINCIPAIS DISPONÍVEIS"], color="bright_green")
    for idx, lista in enumerate(listas_principais, start=1):
        id_colored = color_text(f"(ID:{lista['id']})", **STYLE["dim"])
        print(f"{color_text(str(idx).rjust(3), **STYLE['number'])}. {color_text(lista['nome'], color='bright_white', style='bright')} {id_colored}")
    print(color_text("-" * 80, **STYLE["dim"]))

    while True:
        escolha = input(color_text("Para qual lista do banco principal você deseja migrar? (número): ", **STYLE["info"])).strip()
        if escolha.isdigit():
            idx = int(escolha) - 1
            if 0 <= idx < len(listas_principais):
                main_list = listas_principais[idx]
                break
        print_error("Opção inválida. Tente novamente.")

    # 7. Confirmar
    fancy_header(["📋 RESUMO DA MIGRAÇÃO SELETIVA"], color="bright_cyan")
    print(color_text(f"📌 Lista de espera: {wait_list['nome']}", color="bright_white", style="bright"))
    print(color_text(f"📌 Linhas a migrar: {', '.join(l['nome'] for l in linhas_selecionadas)}", **STYLE["info"]))
    print(color_text(f"📌 Lista destino: {main_list['nome']} (ID {main_list['id']})", color="bright_green", style="bright"))
    print(color_text("=" * 80, **STYLE["dim"]))
    confirm = input(color_text("Confirmar migração? (s/N): ", **STYLE["highlight"])).strip().lower()
    if confirm != 's':
        print_info("Migração cancelada.")
        return

    # 8. Chamar o endpoint seletivo
    url = f"{API_BASE.rstrip('/')}/migrate/wait/to/main/selective"
    payload = {
        "wait_list_id": wait_list_id,
        "linha_ids": [l['id'] for l in linhas_selecionadas],
        "main_list_id": main_list['id']
    }
    try:
        print(color_text("\n⏳ Migrando...", **STYLE["info"]))
        r = requests.post(url, json=payload, timeout=60)
        if r.status_code != 200:
            print_error(f"Erro na migração: {r.status_code} - {r.text}")
            return
        data = r.json()
        print_success(f"Migrados com sucesso: {data.get('migrados', 0)} itens.")
        if data.get('erros'):
            print_warning("Erros:")
            for erro in data['erros']:
                print(color_text(f"  - {erro}", **STYLE["dim"]))
    except Exception as e:
        print_error(f"Erro de rede: {e}")

def cmd_move():
    """
    Comando interativo para mover itens entre listas do banco principal.
    O usuário escolhe os itens pelo número de ordem (posição) exibido na lista.
    """
    listas, err = fetch_lists_request()
    if err:
        print_error(f"Erro ao buscar listas: {err}")
        return
    if not listas:
        print_info("Nenhuma lista principal disponível.")
        return

    fancy_header(["📋 LISTAS PRINCIPAIS DISPONÍVEIS"], color="bright_green")
    for idx, lista in enumerate(listas, start=1):
        id_colored = color_text(f"(ID:{lista['id']})", **STYLE["dim"])
        print(f"{color_text(str(idx).rjust(3), **STYLE['number'])}. {color_text(lista['nome'], color='bright_white', style='bright')} {id_colored}")
    print(color_text("-" * 80, **STYLE["dim"]))

    while True:
        escolha = input(color_text("Qual lista você quer usar como ORIGEM? (número): ", **STYLE["info"])).strip()
        if escolha.isdigit():
            idx = int(escolha) - 1
            if 0 <= idx < len(listas):
                origem = listas[idx]
                break
        print_error("Opção inválida. Tente novamente.")

    linhas, err = fetch_lines_request(origem['id'])
    if err:
        print_error(f"Erro ao buscar linhas: {err}")
        return
    if not linhas:
        print_info("Esta lista está vazia.")
        return

    linhas_ordenadas = sorted(linhas, key=lambda x: x.get('nome', '').casefold())
    fancy_header([f"📋 LISTA ORIGEM: {origem['nome']} (ID {origem['id']})"], color="bright_cyan")
    for idx, linha in enumerate(linhas_ordenadas, start=1):
        id_colored = color_text(f"(ID:{linha['id']})", **STYLE["dim"])
        print(f"{color_text(str(idx).rjust(3), **STYLE['number'])}. {color_text(linha['nome'], color='bright_white', style='bright')} {id_colored}")
    print(color_text("-" * 80, **STYLE["dim"]))

    while True:
        raw = input(color_text("Quais itens você quer mover? (digite os números de posição separados por vírgula, ex: 1,3,5): ", **STYLE["info"])).strip()
        if not raw:
            print_info("Nenhum número informado. Operação cancelada.")
            return
        indices = []
        for part in raw.split(','):
            part = part.strip()
            if part.isdigit():
                indices.append(int(part))
        if not indices:
            print_error("Números inválidos. Tente novamente.")
            continue
        invalidos = [i for i in indices if i < 1 or i > len(linhas_ordenadas)]
        if invalidos:
            print_error(f"Números fora do intervalo (1..{len(linhas_ordenadas)}): {', '.join(map(str, invalidos))}")
            continue
        ids_selecionados = [linhas_ordenadas[i-1]['id'] for i in indices]
        break

    listas_destino = [l for l in listas if l['id'] != origem['id']]
    if not listas_destino:
        print_info("Não há outra lista para mover. Operação cancelada.")
        return

    fancy_header(["📋 LISTAS DESTINO DISPONÍVEIS (excluindo origem)"], color="bright_green")
    for idx, lista in enumerate(listas_destino, start=1):
        id_colored = color_text(f"(ID:{lista['id']})", **STYLE["dim"])
        print(f"{color_text(str(idx).rjust(3), **STYLE['number'])}. {color_text(lista['nome'], color='bright_white', style='bright')} {id_colored}")
    print(color_text("-" * 80, **STYLE["dim"]))

    while True:
        escolha = input(color_text("Para qual lista você deseja mover? (número): ", **STYLE["info"])).strip()
        if escolha.isdigit():
            idx = int(escolha) - 1
            if 0 <= idx < len(listas_destino):
                destino = listas_destino[idx]
                break
        print_error("Opção inválida. Tente novamente.")

    fancy_header(["📋 RESUMO DA MOVIMENTAÇÃO"], color="bright_cyan")
    print(color_text(f"📌 Origem: {origem['nome']} (ID {origem['id']})", color="bright_white", style="bright"))
    print(color_text(f"📌 Itens a mover (posições): {', '.join(str(i) for i in indices)}", **STYLE["info"]))
    print(color_text(f"📌 Destino: {destino['nome']} (ID {destino['id']})", color="bright_green", style="bright"))
    print(color_text("=" * 80, **STYLE["dim"]))
    confirm = input(color_text("Confirmar movimentação? (s/N): ", **STYLE["highlight"])).strip().lower()
    if confirm != 's':
        print_info("Movimentação cancelada.")
        return

    url = f"{API_BASE.rstrip('/')}/move/items"
    payload = {
        "origem_lista_id": origem['id'],
        "destino_lista_id": destino['id'],
        "item_ids": ids_selecionados
    }
    try:
        print(color_text("\n⏳ Movendo...", **STYLE["info"]))
        r = requests.post(url, json=payload, timeout=60)
        if r.status_code != 200:
            print_error(f"Erro na movimentação: {r.status_code} - {r.text}")
            return
        data = r.json()
        print_success(f"Movidos com sucesso: {data.get('movidos', 0)} itens.")
        if data.get('erros'):
            print_warning("Erros:")
            for erro in data['erros']:
                print(color_text(f"  - {erro}", **STYLE["dim"]))
    except Exception as e:
        print_error(f"Erro de rede: {e}")

def print_welcome_banner():
    ascii_art = r"""
   ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄▄  ▄▄▄▄▄▄▄▄▄▄▄ 
  ▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌
  ▐░█▀▀▀▀▀▀▀▀▀ ▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀▀▀▀█░▌▐░█▀▀▀▀▀▀▀▀▀ ▐░█▀▀▀▀▀▀▀▀▀ 
  ▐░▌          ▐░▌       ▐░▌▐░▌       ▐░▌▐░▌          ▐░▌          
  ▐░█▄▄▄▄▄▄▄▄▄ ▐░▌       ▐░▌▐░▌       ▐░▌▐░▌          ▐░▌          
  ▐░░░░░░░░░░░▌▐░▌       ▐░▌▐░▌       ▐░▌▐░▌          ▐░▌          
   ▀▀▀▀▀▀▀▀▀█░▌▐░▌       ▐░▌▐░▌       ▐░▌▐░▌          ▐░▌          
            ▐░▌▐░▌       ▐░▌▐░▌       ▐░▌▐░▌          ▐░▌          
   ▄▄▄▄▄▄▄▄▄█░▌▐░█▄▄▄▄▄▄▄█░▌▐░█▄▄▄▄▄▄▄█░▌▐░█▄▄▄▄▄▄▄▄▄ ▐░█▄▄▄▄▄▄▄▄▄ 
  ▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌▐░░░░░░░░░░░▌
   ▀▀▀▀▀▀▀▀▀▀▀  ▀▀▀▀▀▀▀▀▀▀▀  ▀▀▀▀▀▀▀▀▀▀▀  ▀▀▀▀▀▀▀▀▀▀▀  ▀▀▀▀▀▀▀▀▀▀▀ 
    """
    lines = ascii_art.split('\n')
    for line in lines:
        if line.strip():
            print(color_text(line, color="bright_cyan", style="bright"))
    print(color_text("⚡ Bem-vindo ao List-IT CLI - Gerenciador de Listas Interativo ⚡", color="bright_yellow", style="bright"))
    print(color_text("   Digite 'help' para ver os comandos disponíveis.", **STYLE["dim"]))
    print()

def cmd_clear_wait():
    confirm = input(color_text("⚠️ Tem certeza que deseja limpar todo o banco de espera? (y/N): ", **STYLE["warning"])).strip().lower()
    if confirm == "y":
        url = f"{API_BASE.rstrip('/')}/wait/clear?confirm=true"
        try:
            r = requests.delete(url, timeout=10)
            if r.status_code == 200:
                print_success("Banco de espera limpo com sucesso.")
            else:
                print_error(f"Erro: {r.status_code}")
        except Exception as e:
            print_error(f"Erro: {e}")
    else:
        print_info("Operação cancelada.")

# DEPOIS - Parser mais robusto

def parse_command(line):
    """
    Parseia o comando de forma mais robusta.
    Suporta comandos com aspas simples e duplas.
    """
    line = line.strip()
    if not line:
        return None, []
    
    parts = line.split(' ', 1)
    cmd = parts[0].lower()
    
    if len(parts) == 1:
        return cmd, []
    
    args_str = parts[1]
    args = []
    current = ''
    in_quotes = False
    quote_char = ''
    i = 0
    
    while i < len(args_str):
        ch = args_str[i]
        
        if not in_quotes:
            if ch == '"' or ch == "'":
                in_quotes = True
                quote_char = ch
                i += 1
                continue
            elif ch == ' ':
                if current:
                    args.append(current)
                    current = ''
                i += 1
                continue
            else:
                current += ch
                i += 1
        else:
            if ch == quote_char:
                in_quotes = False
                quote_char = ''
                i += 1
                continue
            else:
                current += ch
                i += 1
    
    if current:
        args.append(current)
    
    return cmd, args

# -------------------------
# Loop principal
# -------------------------
def main():
    
    current_ctx = None
    print_welcome_banner()
    while True:
        try:
            if current_ctx:
                prompt = color_text(f"┌─[{current_ctx.name}]─", **STYLE["header"]) + color_text("$ ", **STYLE["command"])
            else:
                prompt = PROMPT_MAIN
            line = input(prompt).strip()
            if not line:
                continue
            cmd, args = parse_command(line)

            if cmd in ("help", "?"):
                if current_ctx is None:
                    print_help_main()
                elif isinstance(current_ctx, ItemContext):
                    print_help_item()
                else:
                    print_help_list()
                continue

            # --- Comandos globais ---
            if not current_ctx:
                if cmd == "show_lists":
                    (listas, err) = with_minimum_spinner(lambda: fetch_lists_request(), text="Buscando listas...", min_seconds=0.6)
                    fancy_header(["📋 LISTAS DISPONÍVEIS"], color="bright_cyan")
                    if err:
                        print_error(f"Erro: {err}")
                    else:
                        display = PaginatedDisplay(listas or [], "LISTAS DISPONÍVEIS", items_per_page=None)
                        display.render_page()
                    continue

                if cmd == "show_wait_lists":
                    (listas, err) = with_minimum_spinner(lambda: fetch_wait_lists_request(), text="Buscando listas de espera...", min_seconds=0.6)
                    fancy_header(["📋 LISTAS DE ESPERA"], color="bright_magenta")
                    if err:
                        print_error(f"Erro: {err}")
                    else:
                        display = PaginatedDisplay(listas or [], "LISTAS DE ESPERA", items_per_page=None)
                        display.render_page()
                    continue

                if cmd == "create_list":
                    if not args:
                        print_error("Uso: create_list <nome>")
                        continue
                    nome = " ".join(args)
                    cmd_create_list(nome, is_waiting=False)
                    continue

                if cmd == "create_wait_list":
                    if not args:
                        print_error("Uso: create_wait_list <nome>")
                        continue
                    nome = " ".join(args)
                    cmd_create_list(nome, is_waiting=True)
                    continue

                if cmd == "delete_list":
                    if not args:
                        print_error("Uso: delete_list <id|nome>")
                        continue
                    key = " ".join(args)
                    cmd_delete_list(key)
                    continue

                if cmd == "open":
                    if not args:
                        print_error("Uso: open <id|nome>")
                        continue
                    listkey = " ".join(args)
                    ctx = cmd_open_list(listkey, is_waiting=False)
                    if ctx:
                        current_ctx = ctx
                    continue

                if cmd == "open_wait":
                    if not args:
                        print_error("Uso: open_wait <id|nome>")
                        continue
                    listkey = " ".join(args)
                    ctx = cmd_open_list(listkey, is_waiting=True)
                    if ctx:
                        current_ctx = ctx
                    continue

                if cmd == "verify_api":
                    verify_anilist_api()
                    continue

                if cmd == "migrate_wait":
                    lista_id = args[0] if args else None
                    cmd_migrate_wait(lista_id)
                    continue

                if cmd == "move":
                    cmd_move()
                    continue

                if cmd == "clear_wait":
                    cmd_clear_wait()
                    continue

                if cmd in ("clear", "cls"):
                    clear_screen()
                    continue

                if cmd in ("exit", "quit"):
                    print(color_text("👋 Saindo... Até logo!", **STYLE["dim"]))
                    break

                print_error(f"Comando inválido: {cmd}")
                continue

            # --- Contexto de lista ou item ---

            # Comandos de criação de linha (dentro do contexto)
            if cmd == "create_line":
                if not args:
                    print_error("Uso: create_line <nome>")
                    continue
                nome = " ".join(args)
                if isinstance(current_ctx, OpenListContext):
                    current_ctx.interactive_create_line(nome, is_waiting=current_ctx.is_waiting)
                else:
                    print_error("Este comando só pode ser usado dentro de uma lista aberta.")
                continue

            if cmd == "create_wait_line":
                if not args:
                    print_error("Uso: create_wait_line <nome>")
                    continue
                nome = " ".join(args)
                if isinstance(current_ctx, OpenListContext):
                    print_info("Use create_line; a lista atual já define se será criado no banco principal ou de espera.")
                    current_ctx.interactive_create_line(nome, is_waiting=current_ctx.is_waiting)
                else:
                    print_error("Este comando só pode ser usado dentro de uma lista aberta.")
                continue

            # Se estamos dentro de um ItemContext, tratar comandos específicos
            if isinstance(current_ctx, ItemContext):
                # nav
                if cmd in ("next", "n"):
                    nxt = current_ctx.open_adjacent(1)
                    if nxt:
                        current_ctx = nxt
                        current_ctx.show_details()
                    else:
                        print_info("Não há próximo item.")
                    continue
                if cmd in ("prev", "p"):
                    prev = current_ctx.open_adjacent(-1)
                    if prev:
                        current_ctx = prev
                        current_ctx.show_details()
                    else:
                        print_info("Não há item anterior.")
                    continue

                if cmd == "show_details":
                    current_ctx.show_details()
                    continue

                if cmd == "edit" and args:
                    field = args[0]
                    newval = " ".join(args[1:]) if len(args) > 1 else ""
                    if newval == "":
                        print_info("Uso: edit <campo> <novo_valor>  (ou só 'edit' para modo interativo)")
                        continue
                    msg = current_ctx.edit_field(field, newval)
                    typewriter_print(msg, speed=0.003, **STYLE["info"])
                    continue

                if cmd == "edit":
                    msg = current_ctx.interactive_edit()
                    typewriter_print(msg, speed=0.003, **STYLE["info"])
                    continue

                if cmd == "save":
                    ok, msg = current_ctx.save()
                    if ok:
                        print_success(msg)
                        try:
                            pid = current_ctx.item.get("id")
                            parent_items = current_ctx.parent.lines
                            for i, it in enumerate(parent_items):
                                if str(it.get("id")) == str(pid):
                                    parent_items[i] = current_ctx.item
                                    break
                        except Exception:
                            pass
                    else:
                        print_error(f"Erro ao salvar: {msg}")
                    continue

                if cmd == "refresh":
                    ok, msg = current_ctx.refresh()
                    if ok:
                        print_success(msg)
                    else:
                        print_error(f"Erro: {msg}")
                    continue

                if cmd == "delete":
                    confirm = input(color_text("Confirmar exclusão deste item? (y/N): ", **STYLE["warning"])).strip().lower()
                    if confirm == "y":
                        ok, msg = current_ctx.delete()
                        if ok:
                            print_success("Item excluído.")
                            parent = current_ctx.parent
                            parent.fetch_and_cache_lines()
                            current_ctx = parent
                        else:
                            print_error(f"Erro: {msg}")
                    else:
                        print_info("Exclusão cancelada.")
                    continue

                if cmd == "check":
                    ok, msg = current_ctx.check()
                    if ok:
                        print_success(msg)
                    else:
                        print_error(f"Erro: {msg}")
                    continue

                if cmd in ("back", "b"):
                    fancy_header([f"⬅️ Voltando para '{current_ctx.parent.name}'"], color="bright_cyan")
                    current_ctx = current_ctx.parent
                    continue

                # se não tratado, cai no próximo bloco

            # Comandos de lista (OpenListContext)
            if isinstance(current_ctx, OpenListContext):
                # Tratamento de navegação da exibição atual
                if current_ctx.current_display and current_ctx.current_display.handle_command(cmd, args):
                    continue

                if cmd.startswith("sort_"):
                    if not current_ctx.current_display:
                        print_info("Nenhuma exibição ativa para ordenar. Use 'show_lines' primeiro.")
                        continue
                    method = cmd[len("sort_"):]
                    reverse_flag = False
                    if args:
                        if "-r" in args or "--reverse" in args:
                            reverse_flag = True
                    method_key = method
                    if method == "rate" and reverse_flag:
                        method_key = "rate -r"
                    if method_key in ("0-9", "9-0", "a-z", "z-a", "rate", "rate -r"):
                        msg = current_ctx.current_display.apply_sort(method_key)
                        print_info(msg)
                        current_ctx.current_display.render_page(1)
                    else:
                        print_error("Método de sort desconhecido.")
                    continue

                if cmd.startswith("open_"):
                    key = line[len("open_"):].strip()
                    if not key and args:
                        key = " ".join(args)
                    if key.isdigit():
                        item_ctx, err = current_ctx.open_item_by_index(int(key))
                    else:
                        item_ctx, err = current_ctx.open_item_by_name(key)
                    if err:
                        print_error(f"Erro: {err}")
                    else:
                        current_ctx = item_ctx
                        current_ctx.show_details()
                    continue

                # Suporte a 'open <id|nome>' dentro do contexto de lista
                if cmd == "open":
                    if not args:
                        print_info("Uso: open <id|nome>")
                        continue
                    key = " ".join(args)
                    if key.isdigit():
                        item_ctx, err = current_ctx.open_item_by_index(int(key))
                    else:
                        item_ctx, err = current_ctx.open_item_by_name(key)
                    if err:
                        print_error(f"Erro: {err}")
                    else:
                        current_ctx = item_ctx
                        current_ctx.show_details()
                    continue

                if cmd == "show_lines":
                    filtro = " ".join(args) if args else None
                    current_ctx.show_lines(filtro)
                    continue

                if cmd == "show_tags":
                    current_ctx.show_tags()
                    continue

                if cmd.startswith("search_"):
                    termo = cmd[len("search_"):] or (args[0] if args else "")
                    if not termo:
                        print_info("Uso: search_<nome> (ex.: search_Naruto)")
                    else:
                        current_ctx.search_items_by_name(termo)
                    continue

                if cmd.startswith("show_"):
                    resto_comando = cmd[5:]
                    if resto_comando:
                        tags_disponiveis = [tag.lower() for tag in extrair_tags_dos_itens(current_ctx.lines)]
                        tag_correspondente = None
                        for tag in tags_disponiveis:
                            if resto_comando.replace('_', ' ').lower() == tag.lower():
                                tag_correspondente = tag
                                break
                        if tag_correspondente:
                            current_ctx.show_por_tag(tag_correspondente)
                            continue
                    content_commands = {"anime":"anime","filme":"filme","manga":"manga","manhwa":"manhwa","webtoon":"webtoon"}
                    if resto_comando in content_commands:
                        current_ctx.show_por_content(content_commands[resto_comando])
                        continue
                    status_commands = {"seeing":"vendo","finished":"concluido","canceled":"cancelado","see":"assistir","know":"conheço","dropped":"dropado","lendo":"lendo"}
                    if resto_comando in status_commands:
                        current_ctx.show_por_status(status_commands[resto_comando])
                        continue
                    opiniao_commands = {"favorito":"Favorito","muito_bom":"Muito Bom","recomendo":"Recomendo","bom":"Bom","mediano":"Mediano","ruim":"Ruim","horrivel":"Horrível","nao_vi":"Não Vi"}
                    if resto_comando in opiniao_commands:
                        current_ctx.show_por_opiniao(opiniao_commands[resto_comando])
                        continue
                    print_error(f"Comando não reconhecido: {cmd}")
                    continue

                if cmd == "export_list":
                    filename_arg = args[0] if args else None
                    ok, msg = current_ctx.export_current_display(filename_arg)
                    if ok:
                        print_success(msg)
                    else:
                        print_error(f"Erro: {msg}")
                    continue

                if cmd == "next":
                    if current_ctx.current_display:
                        current_ctx.current_display.handle_command("next", [])
                    else:
                        print_info("Nenhuma exibição ativa. Use 'show_lines' primeiro.")
                    continue

                if cmd == "prev":
                    if current_ctx.current_display:
                        current_ctx.current_display.handle_command("prev", [])
                    else:
                        print_info("Nenhuma exibição ativa. Use 'show_lines' primeiro.")
                    continue

                if cmd.isdigit():
                    if current_ctx.current_display:
                        current_ctx.current_display.handle_command(cmd, [])
                    else:
                        print_info("Nenhuma exibição ativa. Use 'show_lines' primeiro.")
                    continue

                if cmd in ("back", "b"):
                    fancy_header([f"⬅️ Saindo do contexto '{current_ctx.name}'"], color="bright_cyan")
                    current_ctx = None
                    continue

                if cmd in ("clear", "cls"):
                    clear_screen()
                    continue

                if cmd in ("exit", "quit"):
                    print(color_text("👋 Saindo... Até logo!", **STYLE["dim"]))
                    break

                print_error(f"Comando inválido em '{current_ctx.name}': {cmd}")

        except (KeyboardInterrupt, EOFError):
            print()
            print(color_text("👋 Saindo... Até logo!", **STYLE["dim"]))
            break

if __name__ == "__main__":
    main()